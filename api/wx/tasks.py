from __future__ import absolute_import, unicode_literals

import hashlib
import json
import logging
import os
import socket
import subprocess
from datetime import datetime, timedelta, timezone
from ftplib import FTP, error_perm, error_reply
from time import sleep, time
from croniter import croniter

import csv
import tempfile
import urllib3
from minio import Minio

import math, cmath
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
import cronex
import dateutil.parser
import pandas
import psycopg2
import pytz
import requests
import subprocess
import wx.export_surface_oscar as exso
from celery import shared_task
from celery.utils.log import get_task_logger
from django.core.cache import cache
from django.db import connection
from django.utils.timezone import now, timedelta
from django.db.models import Count, Q


from tempestas_api import settings
from wx.decoders.flash import read_data as read_data_flash
from wx.decoders.hobo import read_file as read_file_hobo
from wx.decoders.hydro import read_file as read_file_hydrology
from wx.decoders.manual_data import read_file as read_file_manual_data
from wx.decoders.manual_data_hourly import read_file as read_file_manual_data_hourly
from wx.decoders.nesa import read_data as read_data_nesa
from wx.decoders.sat_tx325 import read_data as read_data_sat_tx325
from wx.decoders.surtron import read_data as read_data_surtron
from wx.decoders.surface import read_file as read_file_surface
from wx.decoders.toa5 import read_file as read_file_toa5
from wx.models import DataFile, CombineDataFile
from wx.models import Document
from wx.models import NoaaDcp
from wx.models import Station, Country, WMOReportingStatus, WMORegion, WMOStationType
from wx.models import StationFileIngestion, StationDataFile, ManualStationDataFile
from wx.models import HourlySummaryTask, DailySummaryTask
from wx.models import HydroMLPredictionStation, HydroMLPredictionMapping
from wx.models import HighFrequencyData, HFSummaryTask
from wx.models import LocalWisCredentials, RegionalWisCredentials, Wis2BoxPublishLogs, Wis2BoxPublish
from wx.decoders.insert_raw_data import insert as insert_rd

from django.core.exceptions import ObjectDoesNotExist
import numpy as np
import pandas as pd
from wx.models import Variable
from wx.models import QcPersistThreshold, BackupTask, BackupLog
from wx.enums import QualityFlagEnum

NOT_CHECKED = QualityFlagEnum.NOT_CHECKED.id
SUSPICIOUS = QualityFlagEnum.SUSPICIOUS.id
BAD = QualityFlagEnum.BAD.id
GOOD = QualityFlagEnum.GOOD.id

logger = get_task_logger(__name__)
db_logger = get_task_logger('db')

def get_connection():
    return psycopg2.connect(settings.SURFACE_CONNECTION_STRING)

############################################################


from wx.wave_data_generator import gen_dataframe_and_filename, format_and_send_data
from wx.models import FTPServer

@shared_task
def gen_toa5_file():
    logger.info('Inside dcp_tasks_scheduler')
    df, filename = gen_dataframe_and_filename()
    format_and_send_data(df, filename)

############################################################

def backup_set_running(backup_task, started_at, file_path):
    status = 'RUNNING'
    message = 'Backup process is currently running.'
    new_log = BackupLog(backup_task=backup_task,
                        started_at=started_at,
                        status=status,
                        message=message,
                        file_path=file_path)

    new_log.save()
    return new_log.id

def backup_create(file_path):
    db = settings.DATABASES['default']

    db_host = db['HOST']
    db_port = db['PORT']
    db_name = db['NAME']    
    db_user = db['USER']
    db_pass = db['PASSWORD'] 

    dbname = 'postgresql://'+db_user+':'+db_pass+'@'+db_host+':'+db_port+'/'+db_name

    command = '/usr/bin/pg_dump --dbname=' + dbname + ' | gzip -9 > ' + file_path

    proc = subprocess.Popen(command, shell=True)
    proc.wait()

def backup_ftp(file_name, file_path, ftp_server, remote_folder):
    remote_file_path = os.path.join(remote_folder, file_name)

    with FTP() as ftp:
        ftp.connect(host=ftp_server.host, port=ftp_server.port)
        ftp.login(user=ftp_server.username, passwd=ftp_server.password)
        ftp.set_pasv(val = not ftp_server.is_active_mode)

        with open(file_path, "rb") as file:
            ftp.storbinary(f"STOR {remote_file_path}", file)

        ftp.dir()
        ftp.quit()

def backup_log(log_id, backup_task, started_at, finished_at, status, message, file_path):
    if status in ['SUCCESS', 'FTP ERROR']:
        file_size = os.stat(file_path).st_size/1024 # Bytes -> MegaBytes

        _log = BackupLog.objects.filter(id=log_id)
        _log.update(finished_at=finished_at,
                    status=status,
                    message=message,
                    file_size=file_size)
    else:
        if log_id is None:
            if file_path is None:
                file_name = 'FILE PATH ERROR'

            new_log = BackupLog(backup_task=backup_task,
                                started_at=started_at,
                                finished_at=finished_at,
                                status=status,
                                message=message,
                                file_path=file_path)
            new_log.save()
        else:
            _log = BackupLog.objects.filter(id=log_id)
            _log.update(finished_at=finished_at,
                        status=status,
                        message=message)

def backup_free(backup_task, now):
    delete_datetime = now-timedelta(days=backup_task.retention)
    delete_date = delete_datetime.date()

    _backup_logs = BackupLog.objects.filter(created_at__lt=delete_datetime, backup_task=backup_task)
    file_paths = [_backup_log.file_paths for _backup_log in _backup_logs]

    for file_path in file_paths:
        file_time = os.path.getctime(file_path)
        file_date = datetime.fromtimestamp(file_time).date()
        if file_date < delete_date:
            os.remove(file_path)

    BackupLog.objects.filter(created_at__lt=delete_datetime).delete()

def backup_process(_entry):
    backup_dir = '/data/backup/'

    log_id = None
    file_name = None
    file_path = None    

    started_at = pytz.UTC.localize(datetime.now())
    try:
        file_name = started_at.strftime(_entry.file_name)
        file_path = os.path.join(backup_dir, file_name)

        log_id = backup_set_running(_entry, started_at, file_path)
        backup_create(file_path)
        if _entry.ftp_server is not None:
            try:
                backup_ftp(file_name, file_path, _entry.ftp_server, _entry.remote_folder)
                status = 'SUCCESS'
                message = 'Backup created and successfully send via FTP.'
            except Exception as e:
                status = 'FTP ERROR'
                message = e
                print('Exception happened during backup ftp %s' %(e))                    
        else:
            status = 'SUCCESS'
            message = 'Backup created.'
    except Exception as e:
        status = 'BACKUP ERROR'
        message = e
        print('Exception happened during backup creation %s' %(e))
    finally:
        finished_at = pytz.UTC.localize(datetime.now())

        backup_log(log_id, _entry, started_at, finished_at, status, message, file_path)

        backup_free(_entry, started_at)

@shared_task
def backup_postgres():
    now = pytz.UTC.localize(datetime.now())
    _backup_tasks = BackupTask.objects.filter(is_active=True)
    for _entry in _backup_tasks:
        if croniter.match(_entry.cron_schedule, now):
            backup_process(_entry)

@shared_task
def calculate_hourly_summary(start_datetime=None, end_datetime=None, station_id_list=None):
    start_at = time()

    if not start_datetime:
        start_datetime = datetime.today() - timedelta(hours=48)

    if not end_datetime:
        end_datetime = datetime.today()

    if start_datetime.tzinfo is None:
        start_datetime = pytz.UTC.localize(start_datetime)

    if end_datetime.tzinfo is None:
        end_datetime = pytz.UTC.localize(end_datetime)

    if start_datetime > end_datetime:
        print('Error - start date is more recent than end date.')
        return

    if station_id_list is None:
        station_ids = Station.objects.filter(is_active=True).values_list('id', flat=True)
    else:
        station_ids = station_id_list

    station_ids = tuple(station_ids)

    logger.info('Hourly summary started at {}'.format(datetime.today()))
    logger.info('Hourly summary parameters: {} {} {}'.format(start_datetime, end_datetime, station_id_list))

    delete_sql = f"""
        DELETE FROM hourly_summary 
        WHERE station_id in %(station_ids)s 
          AND datetime = %(start_datetime)s
    """

    insert_sql = f"""
        INSERT INTO hourly_summary (
            datetime,
            station_id,
            variable_id,
            min_value,
            max_value,
            avg_value,
            sum_value,
            num_records,
            created_at,
            updated_at
        )
        SELECT 
            values.datetime,
            values.station_id,
            values.variable_id,
            values.min_value,
            values.max_value,
            values.avg_value,
            values.sum_value,
            values.num_records,
            now(),
            now()
        FROM
            (SELECT 
                CASE WHEN rd.datetime = rd.datetime::date THEN date_trunc('hour', rd.datetime - '1 second'::interval) ELSE date_trunc('hour', rd.datetime) END as datetime,
                station_id,
                variable_id,
                min(calc.value) AS min_value,
                max(calc.value) AS max_value,
                avg(calc.value) AS avg_value,
                sum(calc.value) AS sum_value,
                count(calc.value) AS num_records
            FROM 
                raw_data rd
                ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
            WHERE rd.datetime >= %(start_datetime)s
              AND rd.datetime <= %(end_datetime)s
              AND (rd.manual_flag in (1,4) OR (rd.manual_flag IS NULL AND rd.quality_flag in (1,4)))
              AND NOT rd.is_daily
              AND calc.value != %(MISSING_VALUE)s
              AND station_id in %(station_ids)s
            GROUP BY 1,2,3) values
        WHERE values.datetime = %(start_datetime)s
        UNION ALL
        SELECT 
            values.datetime,
            values.station_id,
            values.variable_id,
            values.min_value,
            values.max_value,
            values.avg_value,
            values.sum_value,
            values.num_records,
            now(),
            now()
        FROM
            (SELECT date_trunc('hour', rd.datetime) as datetime,
                station_id,
                variable_id,
                min(calc.value) AS min_value,
                max(calc.value) AS max_value,
                avg(calc.value) AS avg_value,
                sum(calc.value) AS sum_value,
                count(calc.value) AS num_records
            FROM 
                raw_data rd
                ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
            WHERE rd.datetime >= %(start_datetime)s
              AND rd.datetime <= %(end_datetime)s
              AND (rd.manual_flag in (1,4) OR (rd.manual_flag IS NULL AND rd.quality_flag in (1,4)))
              AND rd.is_daily
              AND calc.value != %(MISSING_VALUE)s
              AND station_id in %(station_ids)s
            GROUP BY 1,2,3) values
        WHERE values.datetime = %(start_datetime)s
    """

    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(delete_sql, {"station_ids": station_ids, "start_datetime": start_datetime})
        cursor.execute(insert_sql, {"station_ids": station_ids, "start_datetime": start_datetime, "end_datetime": end_datetime, "MISSING_VALUE": settings.MISSING_VALUE})
    conn.commit()
    conn.close()
    logger.info(f'Hourly summary finished at {datetime.now(pytz.UTC)}. Took {time() - start_at} seconds.')

@shared_task
def calculate_daily_summary(start_date=None, end_date=None, station_id_list=None):
    logger.info(f'DAILY SUMMARY started at {datetime.now(tz=pytz.UTC)} with parameters: '
                f'start_date={start_date} end_date={end_date} '
                f'station_id_list={station_id_list}')

    start_at = time()

    if start_date is None or end_date is None:
        start_date = datetime.now(pytz.UTC).date()
        end_date = (datetime.now(pytz.UTC) + timedelta(days=1)).date()

    if start_date > end_date:
        print('Error - start_date is more recent than end_date.')
        return

    conn = get_connection()
    with conn.cursor() as cursor:

        if station_id_list is None:
            stations = Station.objects.filter(is_active=True)
        else:
            stations = Station.objects.filter(id__in=station_id_list)

        offsets = list(set([s.utc_offset_minutes for s in stations]))
        for offset in offsets:
            station_ids = tuple(stations.filter(utc_offset_minutes=offset).values_list('id', flat=True))
            fixed_offset = pytz.FixedOffset(offset)

            datetime_start_utc = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0, tzinfo=pytz.UTC)
            datetime_end_utc = datetime(end_date.year, end_date.month, end_date.day, 0, 0, 0, tzinfo=pytz.UTC)

            datetime_start = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0,
                                      tzinfo=fixed_offset).astimezone(pytz.UTC)
            datetime_end = datetime(end_date.year, end_date.month, end_date.day, 0, 0, 0,
                                    tzinfo=fixed_offset).astimezone(pytz.UTC)

            logger.info(f"datetime_start={datetime_start}, datetime_end={datetime_end} "
                        f"offset={offset} "
                        f"station_ids={station_ids}")

            delete_sql = """
                DELETE FROM daily_summary 
                WHERE station_id in %(station_ids)s 
                AND day >= %(datetime_start)s
                AND day < %(datetime_end)s
            """

            insert_sql = """
                INSERT INTO daily_summary (
                    "day",
                    station_id,
                    variable_id,
                    min_value,
                    max_value,
                    avg_value,
                    sum_value,
                    num_records,
                    created_at,
                    updated_at
                )
                SELECT 
                    cast((rd.datetime + interval '%(offset)s minutes') at time zone 'utc' - '1 second'::interval as DATE) as "date",
                    station_id,
                    variable_id,
                    min(calc.value) AS min_value,
                    max(calc.value) AS max_value,
                    avg(calc.value) AS avg_value,
                    sum(calc.value) AS sum_value,
                    count(calc.value) AS num_records,
                    now(),
                    now()
                FROM 
                    raw_data rd
                    ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
                WHERE rd.datetime > %(datetime_start)s
                  AND rd.datetime <= %(datetime_end)s
                  AND calc.value != %(MISSING_VALUE)s
                  AND station_id in %(station_ids)s
                  AND (rd.manual_flag in (1,4) OR (rd.manual_flag IS NULL AND rd.quality_flag in (1,4)))
                  AND NOT rd.is_daily
                GROUP BY 1,2,3
                UNION ALL
                SELECT 
                    cast((rd.datetime + interval '%(offset)s minutes') at time zone 'utc' as DATE) as "date",
                    station_id,
                    variable_id,
                    min(calc.value) AS min_value,
                    max(calc.value) AS max_value,
                    avg(calc.value) AS avg_value,
                    sum(calc.value) AS sum_value,
                    count(calc.value) AS num_records,
                    now(),
                    now()
                FROM 
                    raw_data rd
                    ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
                WHERE rd.datetime > %(datetime_start)s
                  AND rd.datetime <= %(datetime_end)s
                  AND calc.value != %(MISSING_VALUE)s
                  AND station_id in %(station_ids)s
                  AND (rd.manual_flag in (1,4) OR (rd.manual_flag IS NULL AND rd.quality_flag in (1,4)))
                  AND rd.is_daily
                GROUP BY 1,2,3
            """

            cursor.execute(delete_sql, {"datetime_start": datetime_start_utc, "datetime_end": datetime_end_utc,
                                        "station_ids": station_ids})
            cursor.execute(insert_sql,
                           {"datetime_start": datetime_start, "datetime_end": datetime_end, "station_ids": station_ids,
                            "offset": offset, "MISSING_VALUE": settings.MISSING_VALUE})
            conn.commit()


    conn.commit()
    conn.close()

    cache.set('daily_summary_last_run', datetime.today(), None)
    logger.info(f'Daily summary finished at {datetime.now(pytz.UTC)}. Took {time() - start_at} seconds.')

@shared_task
def calculate_station_minimum_interval(start_date=None, end_date=None, station_id_list=None):
    logger.info(f'CALCULATE STATION MINIMUM INTERVAL started at {datetime.now(tz=pytz.UTC)} with parameters: '
                f'start_date={start_date} end_date={end_date} '
                f'station_id_list={station_id_list}')

    start_at = time()

    if start_date is None or end_date is None:
        start_date = datetime.now(pytz.UTC).date()
        end_date = (datetime.now(pytz.UTC) + timedelta(days=1)).date()

    if start_date > end_date:
        print('Error - start_date is more recent than end_date.')
        return

    conn = get_connection()
    with conn.cursor() as cursor:

        if station_id_list is None:
            stations = Station.objects.filter(is_active=True)
        else:
            stations = Station.objects.filter(id__in=station_id_list)

        offsets = list(set([s.utc_offset_minutes for s in stations]))
        for offset in offsets:
            station_ids = tuple(stations.filter(utc_offset_minutes=offset).values_list('id', flat=True))
            fixed_offset = pytz.FixedOffset(offset)

            datetime_start_utc = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0, tzinfo=pytz.UTC)
            datetime_end_utc = datetime(end_date.year, end_date.month, end_date.day, 0, 0, 0, tzinfo=pytz.UTC)

            datetime_start = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0,
                                      tzinfo=fixed_offset).astimezone(pytz.UTC)
            datetime_end = datetime(end_date.year, end_date.month, end_date.day, 0, 0, 0,
                                    tzinfo=fixed_offset).astimezone(pytz.UTC)

            logger.info(f"datetime_start={datetime_start}, datetime_end={datetime_end} "
                        f"offset={offset} "
                        f"station_ids={station_ids}")

            insert_minimum_data_interval = """
                INSERT INTO wx_stationdataminimuminterval (
                     datetime
                    ,station_id
                    ,variable_id
                    ,minimum_interval
                    ,record_count
                    ,ideal_record_count
                    ,record_count_percentage
                    ,created_at
                    ,updated_at
                ) 
                SELECT current_day
                      ,[station_id]station_id
                      ,stationvariable.variable_id
                      ,min(value.data_interval) as minimum_interval
                      ,COALESCE(count(value.formated_datetime), 0) as record_count 
                      ,COALESCE(EXTRACT('EPOCH' FROM interval '1 day') / EXTRACT('EPOCH' FROM min(value.data_interval)), 0) as ideal_record_count
                      ,COALESCE(count(value.formated_datetime) / (EXTRACT('EPOCH' FROM interval '1 day') / EXTRACT('EPOCH' FROM min(value.data_interval))) * 100, 0) as record_count_percentage
                      ,now()
                      ,now()
                FROM generate_series(%(datetime_start)s , %(datetime_end)s , INTERVAL '1 day') as current_day
                    ,wx_stationvariable as stationvariable
                    ,wx_station as station
                LEFT JOIN LATERAL (
                    SELECT date_trunc('day', rd.datetime - INTERVAL '1 second' + (COALESCE(station.utc_offset_minutes, 0)||' minutes')::interval) as formated_datetime
                          ,CASE WHEN rd.is_daily THEN '24:00:00' ELSE LEAD(datetime, 1) OVER (partition by station_id, variable_id order by datetime) - datetime END as data_interval
                    FROM raw_data rd
                    WHERE rd.datetime   > current_day - ((COALESCE(station.utc_offset_minutes, 0)||' minutes')::interval)
                      AND rd.datetime   <= current_day + INTERVAL '1 DAY' - ((COALESCE(station.utc_offset_minutes, 0)||' minutes')::interval)
                      AND rd.station_id  = stationvariable.station_id
                      AND rd.variable_id = stationvariable.variable_id
                ) value ON TRUE
                WHERE stationvariable.station_id IN %(station_ids)s
                  AND stationvariable.station_id = station.id
                  AND (value.formated_datetime = current_day OR value.formated_datetime is null)
                GROUP BY current_day, stationvariable.station_id, stationvariable.variable_id
                  ON CONFLICT (datetime, station_id, variable_id)
                  DO UPDATE SET
                     minimum_interval        = excluded.minimum_interval
                    ,record_count            = excluded.record_count
                    ,ideal_record_count      = excluded.ideal_record_count
                    ,record_count_percentage = excluded.record_count_percentage
                    ,updated_at = now()
            """
            cursor.execute(insert_minimum_data_interval,
                           {"datetime_start": datetime_start_utc, "datetime_end": datetime_end_utc,
                            "station_ids": station_ids})
            conn.commit()

    conn.commit()
    conn.close()

    logger.info(f'Calculate minimum interval finished at {datetime.now(pytz.UTC)}. Took {time() - start_at} seconds.')

@shared_task
def calculate_last24h_summary():
    print('Last 24h summary started at {}'.format(datetime.today()))

    conn = get_connection()

    with conn.cursor() as cursor:
        sql_delete = "DELETE FROM last24h_summary"
        print(sql_delete)
        cursor.execute(sql_delete)

        sql_insert = f"""
            INSERT INTO last24h_summary (
                datetime,
                station_id,
                variable_id,
                min_value,
                max_value,
                avg_value,
                sum_value,
                num_records,
                latest_value
            )
            WITH last AS (
                SELECT
                    datetime,
                    station_id,
                    variable_id,
                    latest_value
                FROM
                    (SELECT 
                        datetime,
                        station_id,
                        variable_id,
                        calc.value AS latest_value,
                        row_number() over (partition by station_id, variable_id order by datetime desc) as rownum
                    FROM 
                        raw_data rd
                        ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
                    WHERE datetime > (now() - interval '1 day')
                    AND datetime <= now()
                    AND (rd.consisted IS NOT NULL OR quality_flag in (1, 4))
                    AND measured != {settings.MISSING_VALUE}
                    AND is_daily = false) AS latest
                WHERE
                    rownum = 1
            ),
            agg AS (
                SELECT 
                    station_id,
                    variable_id,
                    min(calc.value) AS min_value,
                    max(calc.value) AS max_value,
                    avg(calc.value) AS avg_value,
                    sum(calc.value) AS sum_value,
                    count(calc.value) AS num_records
                FROM 
                    raw_data rd
                    ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
                WHERE datetime >  (now() - interval '1 day')
                  AND datetime <= now()
                  AND (rd.consisted IS NOT NULL OR quality_flag in (1, 4))
                  AND calc.value != {settings.MISSING_VALUE}
                  AND is_daily = false
                GROUP BY 1,2
            )
            SELECT
                now(),
                agg.station_id,
                agg.variable_id,
                agg.min_value,
                agg.max_value,
                agg.avg_value,
                agg.sum_value,
                agg.num_records,
                last.latest_value
            FROM
                agg
                JOIN last ON agg.station_id = last.station_id AND agg.variable_id = last.variable_id
            ON CONFLICT (station_id, variable_id) DO
            UPDATE SET
                min_value = excluded.min_value,
                max_value = excluded.max_value,
                avg_value = excluded.avg_value,
                sum_value = excluded.sum_value,
                num_records = excluded.num_records,
                latest_value = excluded.latest_value,
                datetime = excluded.datetime;
        """
        print(sql_insert)
        cursor.execute(sql_insert)

    conn.commit()
    conn.close()

    cache.set('last24h_summary_last_run', datetime.today(), None)
    print('Last 24h summary finished at {}'.format(datetime.today()))

@shared_task
def process_document():
    available_decoders = {
        'HOBO': read_file_hobo,
        'TOA5': read_file_toa5,
        'HYDROLOGY': read_file_hydrology
        # Nesa
    }

    default_decoder = 'TOA5'

    document_list = Document.objects.select_related('decoder', 'station').filter(processed=False).order_by('id')[:60]

    logger.info('Documents: %s' % document_list)

    for document in document_list:
        if document.decoder:
            current_decoder = available_decoders[document.decoder.name]
        else:
            current_decoder = available_decoders[default_decoder]

        logger.info('Processing file "{0}" with "{1}" decoder.'.format(document.file.path, current_decoder))

        try:
            current_decoder(document.file.path, document.station)
        except Exception as err:
            logger.error(
                'Error Processing file "{0}" with "{1}" decoder. '.format(document.file.path, current_decoder) + repr(
                    err))
            db_logger.error(
                'Error Processing file "{0}" with "{1}" decoder. '.format(document.file.path, current_decoder) + repr(
                    err))
        else:
            document.processed = True
            document.save()

@shared_task
def dcp_tasks_scheduler():
    logger.info('Inside dcp_tasks_scheduler')

    noaa_list_to_process = []
    for noaaDcp in NoaaDcp.objects.all():
        if noaaDcp.dcp_address == '50203044':
            DEBUG = True
        else:
            DEBUG = False            

        now = pytz.UTC.localize(datetime.now())

        if noaaDcp.last_datetime is None:
            next_execution = now
        else:
            scheduled_execution = datetime(year=now.year,
                                           month=now.month,
                                           day=now.day,
                                           hour=now.hour,
                                           minute=noaaDcp.first_transmission_time.minute,
                                           second=noaaDcp.first_transmission_time.second)

            transmission_window_timedelta = timedelta(minutes=noaaDcp.transmission_window.minute,
                                                      seconds=noaaDcp.transmission_window.second)

            next_execution = scheduled_execution + transmission_window_timedelta
            next_execution = pytz.UTC.localize(next_execution)

        # if DEBUG or (next_execution <= now and (noaaDcp.last_datetime is None or noaaDcp.last_datetime < next_execution)):
        if next_execution <= now and (noaaDcp.last_datetime is None or noaaDcp.last_datetime < next_execution):
            noaa_list_to_process.append({"noaa_object": noaaDcp, "last_execution": noaaDcp.last_datetime})
            noaaDcp.last_datetime = now
            noaaDcp.save()

    for noaa_dcp in noaa_list_to_process:
        try:
            retrieve_dcp_messages(noaa_dcp)
        except Exception as e:
            logging.error(f'dcp_tasks_scheduler ERROR: {repr(e)}')

def retrieve_dcp_messages(noaa_dict):
    current_noaa_dcp = noaa_dict["noaa_object"]
    last_execution = noaa_dict["last_execution"]

    logger.info('Inside retrieve_dcp_messages ' + current_noaa_dcp.dcp_address)

    related_stations = current_noaa_dcp.noaadcpsstation_set
    related_stations_count = related_stations.count()
    if related_stations_count == 0:
        raise Exception(f"The noaa dcp '{current_noaa_dcp}' is not related to any Station.")
    elif related_stations_count != 1:
        raise Exception(f"The noaa dcp '{current_noaa_dcp}' is related to more than one Station.")

    noaa_dcp_station = related_stations.first()
    station_id = noaa_dcp_station.station_id
    decoder = noaa_dcp_station.decoder.name

    available_decoders = {
        'NESA': read_data_nesa,
        'SAT_TX325': read_data_sat_tx325,  
        # 'SURTRON': read_data_surtron,
    }

    set_search_criteria(current_noaa_dcp, last_execution)

    command = subprocess.Popen([settings.LRGS_EXECUTABLE_PATH,
                                '-h', settings.LRGS_SERVER_HOST,
                                '-p', settings.LRGS_SERVER_PORT,
                                '-u', settings.LRGS_USER,
                                '-P', settings.LRGS_PASSWORD,
                                '-f', settings.LRGS_CS_FILE_PATH
                                ], shell=False, stderr=subprocess.PIPE, stdout=subprocess.PIPE)

    output, err_message = command.communicate()
    response = output.decode('ascii')

    try:
        available_decoders[decoder](station_id, current_noaa_dcp.dcp_address, current_noaa_dcp.config_data, response, err_message)
    except Exception as err:
        logger.error(f'Error on retrieve_dcp_messages for dcp address "{current_noaa_dcp.dcp_address}". {repr(err)}')

def set_search_criteria(dcp, last_execution):
    with open(settings.LRGS_CS_FILE_PATH, 'w') as cs_file:
        if dcp.first_channel is not None:
            cs_file.write(
                f"""DRS_SINCE: now - {dcp_query_window(last_execution)} hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp.dcp_address}\nCHANNEL: |{dcp.first_channel}\n""")
        else:
            cs_file.write(
                f"""DRS_SINCE: now - {dcp_query_window(last_execution)} hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp.dcp_address}\n""")

def dcp_query_window(last_execution):
    return max(3, min(latest_received_dpc_data_in_hours(last_execution), int(settings.LRGS_MAX_INTERVAL)))

def latest_received_dpc_data_in_hours(last_execution):
    try:
        return int(((datetime.now().astimezone(pytz.UTC) - last_execution).total_seconds()) / 3600)
    except TypeError as e:
        return 999999

# set search criteria for dcp testing
def set_search_criteria_dcp_test(dcp_info):
    with open(settings.LRGS_CS_FILE_PATH, 'w') as cs_file:
        if dcp_info['first_channel'] is not None:
            cs_file.write(
                f"""DRS_SINCE: now - 1 hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp_info['dcp_address']}\nCHANNEL: |{dcp_info['first_channel']}\n""")
        else:
            cs_file.write(
                f"""DRS_SINCE: now - 1 hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp_info['dcp_address']}\n""")


# called when the user attempts to test whether a noaaDCP is able to transmit
def test_dcp_transmit(dcp_info):
    logging.info(f"Attempting to test DCP transmition for {dcp_info['dcp_address']}")

    set_search_criteria_dcp_test(dcp_info)

    command = subprocess.Popen([settings.LRGS_EXECUTABLE_PATH,
                                '-h', settings.LRGS_SERVER_HOST,
                                '-p', settings.LRGS_SERVER_PORT,
                                '-u', settings.LRGS_USER,
                                '-P', settings.LRGS_PASSWORD,
                                '-f', settings.LRGS_CS_FILE_PATH
                                ], shell=False, stderr=subprocess.PIPE, stdout=subprocess.PIPE)

    output, err_message = command.communicate()
    # response = {
    #     "output": output.decode('ascii'),
    #     "err_message": err_message
    # }

    response = output.decode('ascii')

    logging.info(f"{dcp_info['dcp_address']} DCP test transmission output {response}")

    return response

@shared_task
def get_entl_data():
    print('LIGHTNING DATA - Starting get_entl_data task...')
    while True:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as entl_socket:
            try:
                print("LIGHTNING DATA - Connecting to ENTLN server: {}:{}".format(settings.ENTL_PRIMARY_SERVER_HOST,
                                                                                  settings.ENTL_PRIMARY_SERVER_PORT))

                entl_socket.connect((settings.ENTL_PRIMARY_SERVER_HOST, settings.ENTL_PRIMARY_SERVER_PORT))
                entl_socket.settimeout(60)
                entl_socket.setsockopt(socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1)
                entl_socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPIDLE, 1)
                entl_socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPINTVL, 15)
                entl_socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPCNT, 4)

                print("LIGHTNING DATA - Authenticating...")
                auth_string = '{"p":"%s","v":3,"f":2,"t":1}' % settings.ENTL_PARTNER_ID
                entl_socket.send(auth_string.encode('latin-1'))
                print("LIGHTNING DATA - Authenticated")

                process_received_data(entl_socket)

            except Exception as e:
                print("LIGHTNING DATA - An error occurred: " + repr(
                    e) + "\nLIGHTNING DATA - Reconnecting in 3 seconds...")
                sleep(3)
        print('LIGHTNING DATA - Connection error. Reconnecting in 15 seconds...')
        sleep(15)

def process_received_data(entl_socket):
    while True:
        data = entl_socket.recv(56)
        if not data:
            return

        if data[1] == '9':
            print("LIGHTNING DATA - Keep alive packet")
        else:
            latitude = int.from_bytes(data[10:14], byteorder='big', signed=True) / 1e7
            longitude = int.from_bytes(data[14:18], byteorder='big', signed=True) / 1e7
            if (15 <= latitude <= 19 and -90 <= longitude <= -87):
                save_flash_data.delay(data.decode('latin-1'))

# Used to save Flash data asynchronously
@shared_task
def save_flash_data(data_string):
    read_data_flash(data_string.encode('latin-1'))

@shared_task
def export_data(station_id, source, start_date, end_date, variable_ids, file_id, agg, displayUTC):
    logger.info(f'Exporting data (file "{file_id}")')

    timezone_offset = pytz.timezone(settings.TIMEZONE_NAME)
    start_date_utc = pytz.UTC.localize(datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'))
    end_date_utc = pytz.UTC.localize(datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S'))

    station = Station.objects.get(pk=station_id)
    current_datafile = DataFile.objects.get(pk=file_id)

    variable_ids = tuple(variable_ids)
    # variable_ids = ','.join([str(x) for x in variable_ids])

    # Diferent data sources have diferents columns names for the measurement data and diferents intervals
    if source == 'raw_data':
        datetime_variable = 'datetime'
        data_source_description = 'Raw data'
        converted_start_date = start_date_utc
        converted_end_date = end_date_utc

    else:
        # measured_source = '''
        #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value
        #          WHEN var.sampling_operation_id = 3      THEN data.min_value
        #          WHEN var.sampling_operation_id = 4      THEN data.max_value
        #          WHEN var.sampling_operation_id = 6      THEN data.sum_value
        #     ELSE data.sum_value END as value '''
        if source == 'hourly_summary':
            datetime_variable = 'datetime'
            date_source = f"(datetime + interval '{station.utc_offset_minutes} minutes') at time zone 'utc' as date"
            data_source_description = 'Hourly summary'
            converted_start_date = start_date_utc
            converted_end_date = end_date_utc

        elif source == 'daily_summary':
            datetime_variable = 'day'
            data_source_description = 'Daily summary'
            date_source = "day::date"
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        elif source == 'monthly_summary':
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
            #         WHEN var.sampling_operation_id = 3      THEN data.min_value
            #         WHEN var.sampling_operation_id = 4      THEN data.max_value
            #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            datetime_variable = 'date'
            date_source = "date::date"
            data_source_description = 'Monthly summary'
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        elif source == 'yearly_summary':
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
            #         WHEN var.sampling_operation_id = 3      THEN data.min_value
            #         WHEN var.sampling_operation_id = 4      THEN data.max_value
            #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            datetime_variable = 'date'
            date_source = "date::date"
            data_source_description = 'Yearly summary'
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

    try:
        variable_dict = {}
        variable_names_string = ''

        with connection.cursor() as cursor_variable:
            cursor_variable.execute(f'''
                SELECT var.symbol
                    ,var.id
                    ,CASE WHEN unit.symbol IS NOT NULL THEN CONCAT(var.symbol, ' - ', var.name, ' (', unit.symbol, ')') 
                        ELSE CONCAT(var.symbol, ' - ', var.name) END as var_name
                FROM wx_variable var 
                LEFT JOIN wx_unit unit ON var.unit_id = unit.id 
                WHERE var.id IN %s
                ORDER BY var.name
            ''', (variable_ids,))

            rows = cursor_variable.fetchall()
            for row in rows:
                variable_dict[row[1]] = row[0]
                variable_names_string += f'{row[2]}   '

        # Iterate over the start and end date day by day to split the queries
        datetime_list = [converted_start_date]
        current_datetime = converted_start_date
        while current_datetime < converted_end_date and (current_datetime + timedelta(days=1)) < converted_end_date:
            current_datetime = current_datetime + timedelta(days=1)
            datetime_list.append(current_datetime)
        datetime_list.append(converted_end_date)

        # adding an extra day to the list when the source is monthly_summary or daily_summary
        # This will cause the last month/day to be included in the output
        if source in ['monthly_summary', 'daily_summary']:
            datetime_list.append(converted_end_date + timedelta(days=1))

            # in the case of querying for the monthly_summary of only 1 month
            # strip the datetime_list of the last value so that there are no duplicates
            if len(datetime_list) == 3 and source == 'monthly_summary':
                datetime_list.pop()
        # adding an extra hour to the list when the souce is hourly_summary
        elif source == 'hourly_summary':
            datetime_list.append(converted_end_date + timedelta(hours=1))
        # adding extra seconds to the list when the souce is raw_data
        # the number of seconds added is based on the data interval selected
        elif source == 'raw_data':
            datetime_list.append(converted_end_date + timedelta(seconds=current_datafile.interval_in_seconds))


        query_result = []
        for i in range(0, len(datetime_list) - 1):
            current_start_datetime = datetime_list[i]
            current_end_datetime = datetime_list[i + 1]

            with connection.cursor() as cursor:
                if source == 'raw_data': 

                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_raw_data = '''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_raw_data = '''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': current_datafile.interval_in_seconds})


                    cursor.execute(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': current_datafile.interval_in_seconds})

                elif source == 'hourly_summary':
                    
                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})
                    
                elif source == 'daily_summary':
                    query_daily = '''
                        WITH processed_data AS (
                            SELECT day ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM daily_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.day >= %(start_datetime)s
                            AND data.day <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 days') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON day = generated_time AND variable.id = variable_id
                    '''

                    logging.info(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                elif source == 'monthly_summary':
                    query_monthly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM monthly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 months') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                elif source == 'yearly_summary':
                    query_yearly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM yearly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date < %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 years') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                query_result = query_result + cursor.fetchall()



        filepath = f'{settings.EXPORTED_DATA_CELERY_PATH}{file_id}.csv'
        date_of_completion = datetime.utcnow()

        with open(filepath, 'w') as f: 

            # modify the displayed start and end date in the csv file based on the summary type
            # show the hour, minute, second, year, month, day in the output
            if data_source_description in ["Raw data", "Hourly summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')
            # show just the year, month and day
            elif data_source_description in ["Daily summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d')
            # show just the year and month
            elif data_source_description in ["Monthly summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m')
            # show just the year
            elif data_source_description in ["Yearly summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y')
                end_date_header = int(end_date_utc.astimezone(timezone_offset).strftime('%Y')) - 1
            # generic: show everything (hour, minute, second, year, month, day)
            else:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')

            f.write(f'Station:{station.code} - {station.name}\n')
            f.write(f'Data source:{data_source_description}\n')
            f.write(f'Description:{variable_names_string}\n')
            f.write(f'Latitude:{station.latitude}\n')
            f.write(f'Longitude:{station.longitude}\n')
            f.write(f'Date of completion:{date_of_completion.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}\n')
            f.write(f'Prepared by:{current_datafile.prepared_by}\n')
            f.write(f'Start date:{start_date_header}, End date:{end_date_header}\n\n')
            if displayUTC:
                f.write('Dates are displayed in UTC\n')
                f.write(f'Start date in UTC:{converted_start_date.strftime("%Y-%m-%d %H:%M:%S")}, End date in UTC:{converted_end_date.strftime("%Y-%m-%d %H:%M:%S")}\n\n')

            # Check the value of the agg to inform aggregation
            if agg == "min":
                # Write the text for "min" to the file
                f.write(f'Aggregation: Minimum\n\n')
            elif agg == "max":
                # Write the text for "max" to the file
                f.write(f'Aggregation: Maximum\n\n')
            elif agg == "sum":
                # Write the text for "sum" to the file
                f.write(f'Aggregation: Sum\n\n')
            elif agg == "avg":
                # Write the text for "avg" to the file
                f.write(f'Aggregation: Average\n\n')
            else:
                # Handle unexpected values of 'agg'
                f.write(f'Aggregation: Instantaneous Raw Data\n\n')


        lines = 0
        if query_result:
            df = pandas.DataFrame(data=query_result).pivot(index=0, columns=1)
            df.rename(columns=variable_dict, inplace=True)
            df.columns = df.columns.droplevel(0)
            if source == 'daily_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                cols = df.columns.tolist()
                cols = cols[-3:] + cols[:-3]
                df = df[cols]
                df = df.drop_duplicates(subset=['Day', 'Month', 'Year'], keep='first')
            elif source == 'monthly_summary':                
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                cols = df.columns.tolist()
                cols = cols[-2:] + cols[:-2]
                df = df[cols]
                df = df.drop_duplicates(subset=['Month', 'Year'], keep='first')
            elif source == 'yearly_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                cols = df.columns.tolist()
                cols = cols[-1:] + cols[:-1]
                df = df[cols]
                df = df.drop_duplicates(subset=['Year'], keep='first')
            else:
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                df['Time'] = df.index.map(lambda x: x.strftime('%H:%M:%S'))
                cols = df.columns.tolist()
                cols = cols[-4:] + cols[:-4]
                df = df[cols]
            df.to_csv(filepath, index=False, mode='a', header=True)
            lines = len(df.index)

        current_datafile.ready = True
        current_datafile.ready_at = date_of_completion
        current_datafile.lines = lines
        current_datafile.save()
        logger.info(f'Data exported successfully (file "{file_id}")')

        # create a new .xlsx version of the file
        convert_csv_xlsx(filepath, station.name, file_id)


    except Exception as e:
        current_datafile.ready = False
        current_datafile.ready_at = datetime.utcnow()
        current_datafile.lines = 0
        current_datafile.save()
        logger.error(f'Error on export data file "{file_id}". {repr(e)}')


# converts csv files to .xlsx format
def convert_csv_xlsx(file_path, file_station, file_id):
    if os.path.exists(file_path) and file_station:
        # open the csv and read all lines
        with open(file_path, 'r') as file:
            lines = file.readlines()

        try:
            # represents the workbook
            wb = Workbook()
            # currently active default worksheet in the workbook
            ws = wb.active
            # Naming the worksheet
            ws.title = f'{file_station}'

            # Write metadata
            for i, line in enumerate(lines):
                # Stop adding metadata once the data table starts
                if "Year" in line:
                    start_data_index = i + 1
                    break
                # writing the metadata to a single column
                ws.cell(row=i + 1, column=1, value=line.strip())

            # Write tabular data headers and rows
            headers = lines[start_data_index - 1].strip().split(",")
            ws.append(headers)  # Add headers

            # Add the tabular data stating from start_data_index till the end of the list
            for line in lines[start_data_index:]:
                row_data = line.strip().split(",")
                ws.append(row_data)
        except Exception as e:
            logger.error(f'An error occured while attempting to convert {file_id}.csv to .xlsx. Error Output: {e}')

        try:
            # Save the workbook
            wb.save(f'{settings.EXPORTED_DATA_CELERY_PATH}{file_id}.xlsx')  
        except Exception as e:
            logger.error(f'An error occured while attempting to save the {file_id}.xlsx file. Error Output: {e}')


# combine .xlsx files into a single .xlsx file
@shared_task
def combine_xlsx_files(station_ids, data_source, start_date, end_date, variable_ids, aggregation, displayUTC, data_interval_seconds, prepared_by, data_source_description, entry_id):
    # grab the current combine xlsx data file entry
    current_datafile = CombineDataFile.objects.get(pk=entry_id)

    try:

        # list containing the data frame of each station
        station_data_frames = []

        for station_id in station_ids:
            # a list to hold the dataframes containing the data queried from the db
            export_data_df = []

            # station dataframe
            station_df = pandas.DataFrame()

            if aggregation:
                for agg in aggregation:

                    export_data_df.append(export_data_xlsx(station_id, data_source, start_date, end_date, variable_ids, agg, displayUTC, data_interval_seconds))
            else:
                export_data_df.append(export_data_xlsx(station_id, data_source, start_date, end_date, variable_ids, aggregation, displayUTC, data_interval_seconds))

            # Getting the columns in common between dataframs pertaining to a specific station
            if data_source in ['raw_data', 'hourly_summary']:
                # selecting the year, month, day, time columns which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :4]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 4:]], axis=1)

            elif data_source == 'daily_summary':
                # selecting the year, month, day columns which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :3]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 3:]], axis=1)

            elif data_source == 'monthly_summary':
                # selecting the year, month columns which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :2]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 2:]], axis=1)

            elif data_source == 'yearly_summary':
                # selecting the year column which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :1]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 1:]], axis=1)

            else:
                raise ValueError("A valid data source was not passed!")


            # add the station dataframe to a list
            station_data_frames.append(station_df)


        # Create a new workbook for the combined file
        combined_workbook = Workbook()
        combined_workbook.remove(combined_workbook.active)  # Remove default sheet

        # the name of the output file will be the celery task id
        output_file = f'{settings.EXPORTED_DATA_CELERY_PATH}combine-{entry_id}.xlsx'

        # retrieving the variable names given then variable ids and converting the names into a string
        variable_dict = {}
        variable_names_string = ''
        try: 
            with connection.cursor() as cursor_variable:
                cursor_variable.execute(f'''
                    SELECT var.symbol
                        ,var.id
                        ,CASE WHEN unit.symbol IS NOT NULL THEN CONCAT(var.symbol, ' - ', var.name, ' (', unit.symbol, ')') 
                            ELSE CONCAT(var.symbol, ' - ', var.name) END as var_name
                    FROM wx_variable var 
                    LEFT JOIN wx_unit unit ON var.unit_id = unit.id 
                    WHERE var.id IN %s
                    ORDER BY var.name
                ''', (tuple(variable_ids),))

                rows = cursor_variable.fetchall()
                for row in rows:
                    variable_dict[row[1]] = row[0]
                    variable_names_string += f'{row[2]}   '
        except Exception as e:
            logger.error(f"AN ERROR OCCURED WITH THE VARIABLE NAMES: {e}")

        # looping through the dataframes of each station
        for x, id in enumerate(station_ids):
            # get the station object
            station = Station.objects.get(pk=id)

            # the number of entries in the data
            lines = len(station_data_frames[x].index)

            # Create a new sheet with the station name
            sheet = combined_workbook.create_sheet(title=f"{station.name} - {station.code}")

            # Write DataFrame to the sheet
            for r_idx, row in enumerate(dataframe_to_rows(station_data_frames[x], index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # insert a new row at the top of the sheet to add the aggregation options and headers
            sheet.insert_rows(1, 12) # insert two rows at row 1

            # raw data does not have any aggregation options
            if data_source != 'raw_data':
                # Get the number of columns in the DataFrame
                num_cols = station_data_frames[x].shape[1]

                # get the amount of cells to merge
                num_merge = len(variable_ids) - 1

                # retrieve how many times a merger should occur
                qty_merge = len(aggregation) - 1

                start_col = num_cols - num_merge # Second-to-last column index (1-based index)
                end_col = num_cols # Last column index

                # Define a bold border style
                bold_border = Border(
                    left=Side(style="thick"),
                    right=Side(style="thick"),
                    top=Side(style="thick"),
                    bottom=Side(style="thick")
                )

                for x in range(qty_merge,-1,-1):
                    # Merge
                    sheet.merge_cells(start_row=12, start_column=start_col, end_row=12, end_column=end_col)
                    cell = sheet.cell(row=12, column=start_col, value=aggregation[x].upper())
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Apply bold border to all merged cells
                    for row in range(12, 13):  # Only row 12
                        for col in range(start_col, end_col + 1):
                            sheet.cell(row=row, column=col).border = bold_border

                    end_col = start_col - 1 # updated end column index
                    start_col = end_col - num_merge  # updated start column index

            date_of_completion = datetime.utcnow()
            timezone_offset = pytz.timezone(settings.TIMEZONE_NAME)

            # add the file headers
            cell = sheet.cell(row=1, column=1, value=f'Station:{station.code} - {station.name}')
            cell = sheet.cell(row=2, column=1, value=f'Data source:{data_source_description}')
            cell = sheet.cell(row=3, column=1, value=f'Description:{variable_names_string}')
            cell = sheet.cell(row=4, column=1, value=f'Latitude:{station.latitude}')
            cell = sheet.cell(row=5, column=1, value=f'Longitude:{station.longitude}')
            cell = sheet.cell(row=6, column=1, value=f'Date of completion:{date_of_completion.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}')
            cell = sheet.cell(row=7, column=1, value=f'Prepared by:{prepared_by}')

            if displayUTC and data_source in ['raw_data','hourly_summary']:
                cell = sheet.cell(row=9, column=1, value=f'Dates are displayed in UTC')
                cell = sheet.cell(row=10, column=1, value=f'Start date:{start_date}, End date:{end_date}')
            else:
                updated_start_date = pytz.UTC.localize(datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'))
                updated_end_date = pytz.UTC.localize(datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S'))
                cell = sheet.cell(row=9, column=1, value=f'Start date:{updated_start_date.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}, End date:{updated_end_date.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}')

        # Save the workbook
        combined_workbook.save(output_file)

        current_datafile.ready = True
        current_datafile.ready_at = date_of_completion
        current_datafile.lines = lines
        current_datafile.save()
        logger.info(f'Combine Data exported successfully (file "combine-{entry_id}")')
    

    except Exception as e:
        current_datafile.ready = False
        current_datafile.ready_at = datetime.utcnow()
        current_datafile.lines = 0
        current_datafile.save()
        logger.error(f'Error on export combine xlsx data file "combine-{entry_id}". Error -  {repr(e)}')      


# returns the data frame for each query ran in order to facilitate combining the files later
def export_data_xlsx(station_id, source, start_date, end_date, variable_ids, agg, displayUTC, data_interval_seconds):

    timezone_offset = pytz.timezone(settings.TIMEZONE_NAME)
    start_date_utc = pytz.UTC.localize(datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'))
    end_date_utc = pytz.UTC.localize(datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S'))

    station = Station.objects.get(pk=station_id)

    variable_ids = tuple(variable_ids)
    # variable_ids = ','.join([str(x) for x in variable_ids])

    # Diferent data sources have diferents columns names for the measurement data and diferents intervals
    if source == 'raw_data':
        datetime_variable = 'datetime'
        data_source_description = 'Raw data'
        converted_start_date = start_date_utc
        converted_end_date = end_date_utc

    else:
        # measured_source = '''
        #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value
        #          WHEN var.sampling_operation_id = 3      THEN data.min_value
        #          WHEN var.sampling_operation_id = 4      THEN data.max_value
        #          WHEN var.sampling_operation_id = 6      THEN data.sum_value
        #     ELSE data.sum_value END as value '''
        if source == 'hourly_summary':
            datetime_variable = 'datetime'
            date_source = f"(datetime + interval '{station.utc_offset_minutes} minutes') at time zone 'utc' as date"
            data_source_description = 'Hourly summary'
            converted_start_date = start_date_utc
            converted_end_date = end_date_utc

        elif source == 'daily_summary':
            datetime_variable = 'day'
            data_source_description = 'Daily summary'
            date_source = "day::date"
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        elif source == 'monthly_summary':
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
            #         WHEN var.sampling_operation_id = 3      THEN data.min_value
            #         WHEN var.sampling_operation_id = 4      THEN data.max_value
            #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            datetime_variable = 'date'
            date_source = "date::date"
            data_source_description = 'Monthly summary'
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        elif source == 'yearly_summary':
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
            #         WHEN var.sampling_operation_id = 3      THEN data.min_value
            #         WHEN var.sampling_operation_id = 4      THEN data.max_value
            #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            datetime_variable = 'date'
            date_source = "date::date"
            data_source_description = 'Yearly summary'
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

    try:
        variable_dict = {}
        variable_names_string = ''

        with connection.cursor() as cursor_variable:
            cursor_variable.execute(f'''
                SELECT var.symbol
                    ,var.id
                    ,CASE WHEN unit.symbol IS NOT NULL THEN CONCAT(var.symbol, ' - ', var.name, ' (', unit.symbol, ')') 
                        ELSE CONCAT(var.symbol, ' - ', var.name) END as var_name
                FROM wx_variable var 
                LEFT JOIN wx_unit unit ON var.unit_id = unit.id 
                WHERE var.id IN %s
                ORDER BY var.name
            ''', (variable_ids,))

            rows = cursor_variable.fetchall()
            for row in rows:
                variable_dict[row[1]] = row[0]
                variable_names_string += f'{row[2]}   '

        # Iterate over the start and end date day by day to split the queries
        datetime_list = [converted_start_date]
        current_datetime = converted_start_date
        while current_datetime < converted_end_date and (current_datetime + timedelta(days=1)) < converted_end_date:
            current_datetime = current_datetime + timedelta(days=1)
            datetime_list.append(current_datetime)
        datetime_list.append(converted_end_date)

        # adding an extra day to the list when the source is monthly_summary or daily_summary
        # This will cause the last month/day to be included in the output
        if source in ['monthly_summary', 'daily_summary']:
            datetime_list.append(converted_end_date + timedelta(days=1))

            # in the case of querying for the monthly_summary of only 1 month
            # strip the datetime_list of the last value so that there are no duplicates
            if len(datetime_list) == 3 and source == 'monthly_summary':
                datetime_list.pop()
        # adding an extra hour to the list when the souce is hourly_summary
        elif source == 'hourly_summary':
            datetime_list.append(converted_end_date + timedelta(hours=1))
        # adding extra seconds to the list when the souce is raw_data
        # the number of seconds added is based on the data interval selected
        elif source == 'raw_data':
            datetime_list.append(converted_end_date + timedelta(seconds=data_interval_seconds))


        query_result = []
        for i in range(0, len(datetime_list) - 1):
            current_start_datetime = datetime_list[i]
            current_end_datetime = datetime_list[i + 1]

            with connection.cursor() as cursor:
                if source == 'raw_data': 

                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_raw_data = '''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_raw_data = '''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': data_interval_seconds})


                    cursor.execute(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': data_interval_seconds})

                elif source == 'hourly_summary':
                    
                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})
                    
                elif source == 'daily_summary':
                    query_daily = '''
                        WITH processed_data AS (
                            SELECT day ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM daily_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.day >= %(start_datetime)s
                            AND data.day <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 days') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON day = generated_time AND variable.id = variable_id
                    '''

                    logging.info(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                elif source == 'monthly_summary':
                    query_monthly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM monthly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 months') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                elif source == 'yearly_summary':
                    query_yearly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM yearly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date < %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 years') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                query_result = query_result + cursor.fetchall()


        if query_result:
            
            df = pandas.DataFrame(data=query_result).pivot(index=0, columns=1)
            df.rename(columns=variable_dict, inplace=True)
            df.columns = df.columns.droplevel(0)
            if source == 'daily_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                cols = df.columns.tolist()
                cols = cols[-3:] + cols[:-3]
                df = df[cols]
                df = df.drop_duplicates(subset=['Day', 'Month', 'Year'], keep='first')
            elif source == 'monthly_summary':                
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                cols = df.columns.tolist()
                cols = cols[-2:] + cols[:-2]
                df = df[cols]
                df = df.drop_duplicates(subset=['Month', 'Year'], keep='first')
            elif source == 'yearly_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                cols = df.columns.tolist()
                cols = cols[-1:] + cols[:-1]
                df = df[cols]
                df = df.drop_duplicates(subset=['Year'], keep='first')
            else:
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                df['Time'] = df.index.map(lambda x: x.strftime('%H:%M:%S'))
                cols = df.columns.tolist()
                cols = cols[-4:] + cols[:-4]
                df = df[cols]

            return df

    except Exception as e:
        logger.error(f'Error on export data as combine xlsx. Error - {repr(e)}')



@shared_task
def ftp_ingest_historical_station_files():
    hist_data = True
    hfreq_data = False   
    ftp_ingest_station_files(hist_data, hfreq_data)

@shared_task
def ftp_ingest_not_historical_station_files():
    hist_data = False
    hfreq_data = False    
    ftp_ingest_station_files(hist_data, hfreq_data)

@shared_task
def ftp_ingest_highfrequency_station_files():
    hist_data = False
    hfreq_data = True
    ftp_ingest_station_files(hist_data, hfreq_data)

def ftp_ingest_station_files(historical_data, highfrequency_data):
    """
    Get and process station data files via FTP protocol

    Parameters: 
        historical_data (bool): flag to process historical station data files

    """
    dt = datetime.now()

    station_file_ingestions = StationFileIngestion.objects.filter(is_active=True, is_historical_data=historical_data, is_highfrequency_data=highfrequency_data)
    station_file_ingestions = [s for s in station_file_ingestions if
                               cronex.CronExpression(s.cron_schedule).check_trigger(
                                   (dt.year, dt.month, dt.day, dt.hour, dt.minute))]

    # List of unique ftp servers
    ftp_servers = list(set([s.ftp_server for s in station_file_ingestions]))

    if highfrequency_data:
        data_type = 'hf_data'
    else:
        data_type = 'raw_data'

    # Loop over connecting to ftp servers, retrieving and processing files
    for ftp_server in ftp_servers:
        logging.info(f'Connecting to {ftp_server}')

        with FTP() as ftp:
            ftp.connect(ftp_server.host, ftp_server.port)
            ftp.login(ftp_server.username, ftp_server.password)
            ftp.set_pasv(not ftp_server.is_active_mode)
            home_folder = ftp.pwd()

            for sfi in [s for s in station_file_ingestions if s.ftp_server == ftp_server]:
                try:
                    ftp.cwd(sfi.remote_folder)
                except error_perm as e:
                    logger.error(f'Error on access the directory "{sfi.remote_folder}". {repr(e)}')
                    db_logger.error(f'Error on access the directory "{sfi.remote_folder}". {repr(e)}')

                # list remote files
                remote_files = ftp.nlst(sfi.file_pattern)

                for fname in remote_files:
                    try:
                        local_folder = '/data/documents/ingest/%s/%s/%s/%04d/%02d/%02d' % (
                            sfi.decoder.name, sfi.station.code, data_type, dt.year, dt.month, dt.day)
                        local_filename = '%04d%02d%02d%02d%02d%02d_%s' % (
                            dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second, fname)
                        local_path = '%s/%s' % (local_folder, local_filename)
                        os.makedirs(local_folder, exist_ok=True)

                        hash_md5 = hashlib.md5()
                        if sfi.is_binary_transfer:
                            with open(local_path, 'wb') as fp_binary:
                                ftp.retrbinary(f'RETR {fname}',
                                               lambda data: [fp_binary.write(data), hash_md5.update(data)])
                        else:
                            with open(local_path, 'w') as fp:
                                ftp.retrlines(f'RETR {fname}', lambda line: [fp.write(line + '\n'),
                                                                             hash_md5.update(line.encode('utf8'))])

                        if sfi.delete_from_server:
                            try:
                                ftp.delete(fname)
                            except error_perm as e:
                                logger.error(
                                    'Permission error on delete the ftp server file "{0}".'.format(local_path) + repr(
                                        e))
                                db_logger.error(
                                    'Permission error on delete the ftp server file "{0}".'.format(local_path) + repr(
                                        e))
                            except error_reply as e:
                                logger.error('Unknown reply received "{0}".'.format(local_path) + repr(e))
                                db_logger.error('Unknown reply received "{0}".'.format(local_path) + repr(e))

                        # Inserts a StationDataFile object with status = 1 (Not processed)
                        station_data_file = StationDataFile(station=sfi.station
                                                            , decoder=sfi.decoder
                                                            , status_id=1
                                                            , utc_offset_minutes=sfi.utc_offset_minutes
                                                            , filepath=local_path
                                                            , file_hash=hash_md5.hexdigest()
                                                            , file_size=os.path.getsize(local_path)
                                                            , is_historical_data=sfi.is_historical_data
                                                            , is_highfrequency_data=sfi.is_highfrequency_data
                                                            , override_data_on_conflict=sfi.override_data_on_conflict)
                        station_data_file.save()
                        logging.info(f'Downloaded FTP file: {local_path}')
                    except OSError as e:
                        logger.error('OS error. ' + repr(e))
                        db_logger.error('OS error. ' + repr(e))

                ftp.cwd(home_folder)

    process_station_data_files(historical_data, highfrequency_data)

def process_station_data_files(historical_data=False, highfrequency_data=False, force_reprocess=False):
    """
    Process station data files

    Parameters: 
        historical_data (bool): flag to process historical files
        force_reprocess (bool): force file to be reprocessed, don't check if the file is already processed
    """

    available_decoders = {
        'HOBO': read_file_hobo,
        'TOA5': read_file_toa5,
        'HYDROLOGY': read_file_hydrology,
        'BELIZE MANUAL DAILY DATA': read_file_manual_data,
        'BELIZE MANUAL HOURLY DATA': read_file_manual_data_hourly,
        'SURFACE': read_file_surface,
    }

    # Get StationDataFile to process
    # Filter status id to process only StationDataFiles with code 1 (Not processed) or 6 (Reprocess)
    station_data_file_list = (StationDataFile.objects.select_related('decoder', 'station')
                                  .filter(status_id__in=(1, 6), is_historical_data=historical_data, is_highfrequency_data=highfrequency_data).order_by('id')[:60])
    logger.info('Station data files: %s' % station_data_file_list)

    # Mark all file as Being processed to avoid reprocess
    for station_data_file in station_data_file_list:
        # Update status id to 2 (Being processed)
        station_data_file.status_id = 2
        station_data_file.save()

    for station_data_file in station_data_file_list:
        # if force_reprocess is true, ignore if file already exist on the database
        if not force_reprocess:
            # Verify if the file was already processed
            # Check if exists some StationDataFilestatus
            # object with the same file_hash and status different than 4 (Error) or 5 (Skipped)
            file_already_processed = (StationDataFile.objects.filter(file_hash=station_data_file.file_hash)
                                      .exclude(id=station_data_file.id).exclude(status_id__in=(4, 5)).exists())
            if file_already_processed:
                # Update status id to 5 (Skipped)
                station_data_file.status_id = 5
                station_data_file.save()
                continue

        try:
            current_decoder = available_decoders[station_data_file.decoder.name]
            logger.info('Processing file "{0}" with "{1}" decoder.'.format(station_data_file.filepath, current_decoder))

            current_decoder(filename=station_data_file.filepath
                            , highfrequency_data=station_data_file.is_highfrequency_data
                            , station_object=station_data_file.station
                            , utc_offset=station_data_file.utc_offset_minutes
                            , override_data_on_conflict=station_data_file.override_data_on_conflict)

        except Exception as err:
            # Update status id to 4 (Error)
            station_data_file.status_id = 4
            station_data_file.observation = ('Error Processing file with "{0}" decoder. '
                                             .format(current_decoder) + repr(err))[:1024]
            station_data_file.save()

            logger.error('Error Processing file "{0}" with "{1}" decoder. '
                         .format(station_data_file.filepath, current_decoder) + repr(err))
            db_logger.error('Error Processing file "{0}" with "{1}" decoder. '
                            .format(station_data_file.filepath, current_decoder) + repr(err))
        else:
            # Update status id to 3 (Processed)
            station_data_file.status_id = 3
            station_data_file.save()


# modify manual station data file object with the Month and Station_List
@shared_task
def ingest_manual_station_files(data_file_id_list):
    """
    Process station data files for Month and Station Names.
    """

    try:

        for entry_id in data_file_id_list:
            manual_station_data_file = ManualStationDataFile.objects.get(id=entry_id)

            filepath = manual_station_data_file.filepath

            logger.info(f'Processing Manual Datafile: {filepath}')

            source = pandas.ExcelFile(filepath)

            stations_list = '       '.join(source.sheet_names) # combining stations list into a long sting

            # column names
            column_names = ['day', 'PRECIP', 'TEMPMAX', 'TEMPMIN', 'TEMPAVG', 'WNDMIL', 'WINDRUN', 'SUNSHNHR', 'EVAPINI', 'EVAPRES', 'EVAPPAN', 'TEMP', 'TEMPWB', 'TSOIL1', 'TSOIL4', 'DYTHND', 'DYFOG', 'DYHAIL', 'DYGAIL', 'TOTRAD', 'RH@TMAX', 'RHMAX', 'RHMIN',]

            # getting the month information from sheet 1
            sheet_raw_data = source.parse(
                source.sheet_names[0],
                skipfooter=2,
                na_filter=False,
                names=column_names,
                usecols='A:W'
            )

            if not sheet_raw_data.empty:
                header = sheet_raw_data[0:1]
                sheet_month = header['WINDRUN'][0].replace('MONTH: ', '').strip()

            
            # update the manual station data file object
            manual_station_data_file.stations_list = stations_list
            manual_station_data_file.month = sheet_month
            manual_station_data_file.save()

            # process manual station data files
            process_manual_station_data_files(data_file_id_list)

    except Exception as e:
        logger.error('OS error. ' + repr(e))
        db_logger.error('OS error. ' + repr(e))


# process manual station data files
def process_manual_station_data_files(data_file_id_list):
    """
    Process manual station data files

    Parameters: 
        data_file_id_list: list containing the id of the station data files to process
    """

    available_decoders = {
        'BELIZE MANUAL DAILY DATA': read_file_manual_data,
    }

    # Get ManualStationDataFile to process
    # Filter status id to process only StationDataFiles with code 1 (Not processed) or 6 (Reprocess)
    station_data_file_list = ManualStationDataFile.objects.filter(id__in=data_file_id_list)

    # Mark all files as Being processed
    for station_data_file in station_data_file_list:
        # Update status id to 2 (Being processed)
        station_data_file.status_id = 2
        station_data_file.save()

        try:
            current_decoder = available_decoders['BELIZE MANUAL DAILY DATA']
            logger.info('Processing file "{0}" with "{1}" decoder.'.format(station_data_file.filepath, current_decoder))

            current_decoder(filename=station_data_file.filepath
                            , override_data_on_conflict=station_data_file.override_data_on_conflict)

        except Exception as err:
            # Update status id to 4 (Error)
            station_data_file.status_id = 4
            station_data_file.observation = (err)
            station_data_file.save()

            logger.error('Error Processing file "{0}" with "{1}" decoder. '
                         .format(station_data_file.filepath, current_decoder) + repr(err))
            db_logger.error('Error Processing file "{0}" with "{1}" decoder. '
                            .format(station_data_file.filepath, current_decoder) + repr(err))
        else:
            # Update status id to 3 (Processed)
            station_data_file.status_id = 3
            station_data_file.save()

        try:
            # attempting to delete already process manual station data file regardless of status
            os.remove(station_data_file.filepath)
            logger.info(f'Manual station data file deleted after processing. File path: {station_data_file.filepath}')
        except Exception as e:
            logger.error('An error occured whill attempting to delete manual station data file.')
            logger.error(f'Error - {repr(e)}')


# Persist Logic Starts here
# Get hourly data from raw data
def get_hourly_raw_data(start_datetime, end_datetime, station_ids):
    con = get_connection()
    sql = '''SELECT *
             FROM raw_data
             WHERE datetime BETWEEN %(start_datetime)s AND %(end_datetime)s
               AND station_id IN %(station_ids)s
               -- AND qc_persist_quality_flag IS NULL
             ORDER BY datetime DESC
          '''
    params = {"station_ids": tuple(station_ids), "start_datetime": start_datetime, "end_datetime": end_datetime}
    

    sql_query = pd.read_sql_query(sql=sql, con=con, params=params)
    con.close()

    df = pd.DataFrame(sql_query)
    return df

# Station and variables used
def get_hourly_sv_dict(start_datetime, end_datetime, station_ids):
    df = get_hourly_raw_data(start_datetime, end_datetime, station_ids)

    dict_sv = {}
    for station_id in station_ids:
        variable_ids = df[df['station_id']==station_id].variable_id.unique()
        dict_v = {}
        for variable_id in variable_ids:
            s_datetime = df[(df['station_id']==station_id) & (df['variable_id']==variable_id)].datetime.min()
            e_datetime = df[(df['station_id']==station_id) & (df['variable_id']==variable_id)].datetime.max()
            dict_v[variable_id] = (s_datetime, e_datetime)
        dict_sv[station_id] = dict_v
    return dict_sv

# Get data from each statation and variable with window interval
def get_hourly_sv_data(start_datetime, end_datetime, station_id, variable_id, window):
    con = get_connection()
    sql = '''SELECT *
             FROM raw_data
             WHERE (datetime BETWEEN %(start_datetime)s AND %(end_datetime)s)
               AND station_id = %(station_id)s
               AND variable_id = %(variable_id)s
          '''
    params = {"station_id": station_id, "variable_id": int(variable_id), "start_datetime": start_datetime-timedelta(seconds=window), "end_datetime": end_datetime}

    sql_query = pd.read_sql_query(sql=sql, con=con, params=params)
    con.close()

    df = pd.DataFrame(sql_query)

    if not df.empty:
        mask = df.datetime.between(start_datetime, end_datetime)
        df['updated'] = False
        df.loc[mask, 'updated'] = True
    return df

def most_frequent(List):
    return max(set(List), key = List.count)

# Interval
def get_interval(df):
    interval_list = df.datetime.diff(periods=1).dt.total_seconds().replace(np.nan, 0).astype("int32")
    interval_list = list(interval_list)
    interval = most_frequent(interval_list)
    return abs(interval)

# Window
def get_window(station_id, variable_id):
    try:
        _station = Station.objects.get(id=station_id)
        _variable = Variable.objects.get(id=variable_id) 
    except ObjectDoesNotExist:
        return 3600
        
    if  _station.is_automatic:
        hours = _variable.persistence_window_hourly
        if hours is None:
            hours = 1 # 1 Hour
    else:
        hours = _variable.persistence_window
        if hours is None:
            hours = 96 # 4 Days
    return hours*3600

# Thresholds
def get_thresholds(station_id, variable_id, interval, window):
    thresholds = {}
    if window == 0 or interval==0:
        return thresholds

    try:
        # Trying to set persist thresholds using current station
        _persist = QcPersistThreshold.objects.get(station_id=station_id, variable_id=variable_id, interval=interval, window=window)
        thresholds['persist_min'] = _persist.minimum_variance
        thresholds['persist_des'] = 'Custom station Threshold'
    except ObjectDoesNotExist:
        try:
            # Trying to set persist thresholds using current station with NULL intervall
            _range = QcPersistThreshold.objects.get(station_id=station_id, variable_id=variable_id, interval__isnull=True, window=window)        
            thresholds['persist_min'] = _persist.minimum_variance
            thresholds['persist_des'] = 'Custom station Threshold'
        except ObjectDoesNotExist:
            try:
                # Trying to set persist thresholds using current station
                _station = Station.objects.get(pk=station_id)
                _persist = QcPersistThreshold.objects.get(station_id=_station.reference_station_id, variable_id=variable_id, interval=interval, window=window)
                thresholds['persist_min'] = _persist.minimum_variance
                thresholds['persist_des'] = 'Reference station threshold'
            except ObjectDoesNotExist:
                try:
                    # Trying to set persist thresholds using current station with NULL intervall
                    _station = Station.objects.get(pk=station_id)
                    _persist = QcPersistThreshold.objects.get(station_id=_station.reference_station_id, variable_id=variable_id, interval__isnull=True, window=window)        
                    thresholds['persist_min'] = _persist.minimum_variance
                    thresholds['persist_des'] = 'Reference station threshold'
                except ObjectDoesNotExist:
                    try:
                        # Trying to set persistence thresholds using global ranges
                        _station = Station.objects.get(pk=station_id)
                        _persist = Variable.objects.get(pk=variable_id)
                        if _station.is_automatic:
                            thresholds['persist_min'] = _persist.persistence_hourly
                            thresholds['persist_des'] = 'Global threshold (Automatic)'
                        else:
                            thresholds['persist_min'] = _persist.persistence
                            thresholds['persist_des'] = 'Global threshold (Manual)'
                    except ObjectDoesNotExist:
                        pass;                    
    return thresholds

# Persistance function and calculation
def persit_function(values):
    return abs(max(values)-min(values))

def get_persist(row, df, window):
    datetime = row.datetime

    s_datetime = datetime-timedelta(seconds=window)
    e_datetime = datetime

    mask = df.datetime.between(s_datetime, e_datetime)
    List = list(df[mask]['measured'])

    persist = persit_function(List)
    return persist

def qc_persist(value, thresholds):
    if 'persist_min' not in thresholds:
        return NOT_CHECKED, "Threshold not found"

    p_min = thresholds['persist_min']
    p_des = thresholds['persist_des']
           
    if p_min is None:
        return NOT_CHECKED, "Threshold not found"           

    if value >= p_min:
        return GOOD, p_des
    else:
        return BAD, p_des    

def set_persist_sus(row, df, window, persist_flag):
    updated = row.updated
    datetime = row.datetime

    if persist_flag in [BAD, SUSPICIOUS]:
        return persist_flag, updated

    s_datetime = datetime
    e_datetime = datetime+timedelta(seconds=window)

    mask = df.datetime.between(s_datetime, e_datetime)
    List = list(df[mask]['qc_persist_quality_flag'])

    if BAD in List:
        return SUSPICIOUS, True
    return persist_flag, updated

def qc_final(row, persist_flag):
    range_flag = row.qc_range_quality_flag
    step_flag = row.qc_step_quality_flag

    flags = (persist_flag, range_flag, step_flag)
    if BAD in flags:
        return BAD
    elif GOOD in flags:
        return GOOD
    elif SUSPICIOUS in flags:
        return SUSPICIOUS        
    return NOT_CHECKED

def set_persist(row, df, s_datetime, e_datetime, interval, window, thresholds):
    if s_datetime <= row.datetime <= e_datetime:
        persist = get_persist(row, df, window)
        persist_flag, persist_des = qc_persist(persist, thresholds)
    else:
        persist_flag, persist_des = row.qc_persist_quality_flag, row.qc_persist_description

    # By the pandas apply function order, the previous rows are calculated so we can set suspicous flags using them
    persist_flag, updated = set_persist_sus(row, df, window, persist_flag)

    # Finally compute the final flag using all three flags, range, step, and persistence
    final_flag =  qc_final(row, persist_flag)

    return updated, persist_flag, persist_des, final_flag

# Persistance update
def update_insert_persist(df):
    data = df.to_dict('records')
    query = '''
            INSERT INTO raw_data(created_at, updated_at, datetime, station_id, variable_id, measured, qc_persist_quality_flag, qc_persist_description, quality_flag)
                VALUES (CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %(datetime)s, %(station_id)s, %(variable_id)s, 0, %(qc_persist_quality_flag)s, %(qc_persist_description)s, %(quality_flag)s)
            ON CONFLICT (datetime, station_id, variable_id) DO UPDATE SET
                updated_at=CURRENT_TIMESTAMP,
                qc_persist_quality_flag = %(qc_persist_quality_flag)s,
                qc_persist_description = %(qc_persist_description)s,
                quality_flag = %(quality_flag)s
            '''
    con = get_connection()            
    with con.cursor() as cursor:
        cursor.executemany(query, data)
    con.commit()
    con.close()

# Main Persist Function
def update_qc_persist(start_datetime, end_datetime, station_ids, summary_type):
    dict_sv = get_hourly_sv_dict(start_datetime, end_datetime, station_ids)
    for station_id in dict_sv:
        for variable_id in dict_sv[station_id]:
            s_datetime, e_datetime = dict_sv[station_id][variable_id]
            window = get_window(station_id, variable_id)
            df = get_hourly_sv_data(s_datetime, e_datetime, station_id, variable_id, window)
            
            if not df.empty:
                interval = get_interval(df)

                thresholds = get_thresholds(station_id, variable_id, interval, window)   
                
                columns = ['updated', 'qc_persist_quality_flag', 'qc_persist_description', 'quality_flag']
                df[columns] = df.apply(lambda row: set_persist(row, df, s_datetime, e_datetime, interval, window, thresholds), axis=1, result_type="expand")

                df = df[df['updated']==True]                

                columns = ['station_id', 'variable_id', 'qc_persist_quality_flag', 'qc_persist_description', 'quality_flag', 'datetime']
                update_insert_persist(df[columns])

                recalculate_summary(df, station_id, s_datetime, e_datetime, summary_type)

def insert_summay(date, station_id, summary_type):
    query_hourly = '''
                    INSERT INTO wx_hourlysummarytask(created_at, updated_at, datetime, started_at, finished_at, station_id)
                       VALUES (CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %(date)s, NULL, NULL, %(station_id)s)
                   '''
    query_daily = '''
                    INSERT INTO wx_dailysummarytask(created_at, updated_at, date, started_at, finished_at, station_id)
                       VALUES (CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %(date)s, NULL, NULL, %(station_id)s)
                   '''

    if summary_type == 'hourly':
        query = query_hourly
    elif summary_type == 'daily':
        query = query_daily


    data = {'station_id': station_id, 'date': date}
    con = get_connection()            
    with con.cursor() as cursor:
        cursor.execute(query, data)
    con.commit()
    con.close()

def recalculate_summary(df, station_id, s_datetime, e_datetime, summary_type):
    mask = df.datetime.between(s_datetime, e_datetime)    
    datetimes = list(df[~mask]['datetime'])

    if datetimes:
        if summary_type == 'hourly':
            datetimes = set([dt.replace(microsecond=0, second=0, minute=0) for dt in datetimes])
            for datetime in datetimes:
                _hourlysummaries = HourlySummaryTask.objects.filter(station_id=station_id, datetime=datetime, started_at__isnull=True)
                if not _hourlysummaries:
                    insert_summay(datetime, station_id, summary_type)

        elif summary_type == 'daily':
            dates = set([dt.date() for dt in datetimes])
            for date in dates:
                _dailysummaries = DailySummaryTask.objects.filter(station_id=station_id, date=date, started_at__isnull=True)
                if not _hourlysummaries:
                    insert_summay(date, station_id, summary_type)

def hourly_summary(hourly_summary_tasks_ids, station_ids, s_datetime, e_datetime):
    try:
        HourlySummaryTask.objects.filter(id__in=hourly_summary_tasks_ids).update(started_at=datetime.now(tz=pytz.UTC))
        calculate_hourly_summary(s_datetime, e_datetime, station_id_list=station_ids)
    except Exception as err:
        logger.error('Error calculation hourly summary for hour "{0}". '.format(hourly_summary_datetime) + repr(err))
        db_logger.error('Error calculation hourly summary for hour "{0}". '.format(hourly_summary_datetime) + repr(err))
    else:
        HourlySummaryTask.objects.filter(id__in=hourly_summary_tasks_ids).update(finished_at=datetime.now(tz=pytz.UTC))

@shared_task
def process_hourly_summary_tasks():
    # Process only 500 hourly summaries per execution
    unprocessed_hourly_summary_datetimes = HourlySummaryTask.objects.filter(started_at=None).values_list('datetime',flat=True).distinct()[:501]
    for hourly_summary_datetime in unprocessed_hourly_summary_datetimes:
        start_datetime = hourly_summary_datetime
        end_datetime = hourly_summary_datetime + timedelta(hours=1)

        hourly_summary_tasks = HourlySummaryTask.objects.filter(started_at=None, datetime=hourly_summary_datetime)
        hourly_summary_tasks_ids = list(hourly_summary_tasks.values_list('id', flat=True))
        station_ids = list(hourly_summary_tasks.values_list('station_id', flat=True).distinct())

        if station_ids:
            update_qc_persist(start_datetime, end_datetime, station_ids, 'hourly')

        # Updating Hourly summary task
        hourly_summary(hourly_summary_tasks_ids, station_ids, start_datetime, end_datetime)

def daily_summary(daily_summary_tasks_ids, station_ids, s_datetime, e_datetime):
    try:
        DailySummaryTask.objects.filter(id__in=daily_summary_tasks_ids).update(started_at=datetime.now(tz=pytz.UTC))
        calculate_daily_summary(s_datetime, e_datetime, station_id_list=station_ids)
        for station_id in station_ids:
           calculate_station_minimum_interval(s_datetime, e_datetime, station_id_list=(station_id,))
    except Exception as err:
        logger.error('Error calculation daily summary for day "{0}". '.format(daily_summary_date) + repr(err))
        db_logger.error('Error calculation daily summary for day "{0}". '.format(daily_summary_date) + repr(err))
    else:
        DailySummaryTask.objects.filter(id__in=daily_summary_tasks_ids).update(finished_at=datetime.now(tz=pytz.UTC))

@shared_task
def process_daily_summary_tasks():
    # process only 500 daily summaries per execution
    unprocessed_daily_summary_dates = DailySummaryTask.objects.filter(started_at=None).values_list('date',flat=True).distinct()[:501]
    for daily_summary_date in unprocessed_daily_summary_dates:

        start_date = daily_summary_date
        end_date = daily_summary_date + timedelta(days=1)

        daily_summary_tasks = DailySummaryTask.objects.filter(started_at=None, date=daily_summary_date)
        daily_summary_tasks_ids = list(daily_summary_tasks.values_list('id', flat=True))
        station_ids = list(daily_summary_tasks.values_list('station_id', flat=True).distinct())

        if station_ids:
            update_qc_persist(start_date, end_date, station_ids, 'daily')

        daily_summary(daily_summary_tasks_ids, station_ids, start_date, end_date)

def predict_data(start_datetime, end_datetime, prediction_id, station_ids, target_station_id, variable_id,
                 data_period_in_minutes, interval_in_minutes, result_mapping):
    data_frequency = (interval_in_minutes // data_period_in_minutes) - 1
    date_dict = {}

    logger.info(
        f"predict_data= start_datetime: {start_datetime}, end_datetime: {end_datetime}, station_ids: {station_ids}, variable_id: {variable_id}, data_period_in_minutes: {data_period_in_minutes}, interval_in_minutes: {interval_in_minutes}, data_frequency: {data_frequency}")
    query = """
        WITH acc_query AS (SELECT datetime
                                ,station_id
                                ,SUM(measured) OVER (PARTITION BY station_id ORDER BY datetime ROWS BETWEEN %(data_frequency)s PRECEDING AND CURRENT ROW) AS acc
                                ,LAG(datetime, %(data_frequency)s) OVER (PARTITION BY station_id ORDER BY datetime) AS earliest_datetime
                           FROM raw_data
                           WHERE station_id in %(station_ids)s
                             AND variable_id = %(variable_id)s
                             AND datetime   >= %(start_datetime)s
                             AND datetime   <= %(end_datetime)s
                             AND measured   != %(MISSING_VALUE)s)
        SELECT acc_query.datetime
              ,acc_query.station_id
              ,acc_query.acc
        FROM acc_query
        WHERE acc_query.datetime - acc_query.earliest_datetime < INTERVAL '%(interval_in_minutes)s MINUTES';
    """

    params = {
        "start_datetime": start_datetime,
        "end_datetime": end_datetime,
        "station_ids": station_ids,
        "variable_id": variable_id,
        "data_frequency": data_frequency,
        "interval_in_minutes": interval_in_minutes,
        "MISSING_VALUE": settings.MISSING_VALUE
    }

    formated_list = []
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(query, params)

        # Group records in a dictionary by datetime
        # rows[0] = datetime
        # rows[1] = station_id
        # rows[2] = acc

        rows = cursor.fetchall()
        if len(rows) == 0:
            raise Exception('No data found')

        for row in rows:
            current_datetime = row[0]
            current_station_id = row[1]
            current_value = row[2]

            if current_datetime not in date_dict:
                date_dict[current_datetime] = {}

            date_dict[current_datetime][current_station_id] = current_value

        # Validate if a datetime contains all stations measurements, calculate avg and format value 
        for datetime, station_data_dict in date_dict.items():

            current_record_station_ids = tuple(station_data_dict.keys())
            if any(station_id not in current_record_station_ids for station_id in station_ids):
                continue

            current_record_dict = {'datetime': datetime.isoformat()}
            station_count = 0
            value_acc = 0.0

            for station_id, measured in station_data_dict.items():
                current_record_dict[station_id] = measured
                station_count += 1
                value_acc += measured

            if station_count > 0:
                current_record_dict['avg'] = value_acc / station_count
                formated_list.append(current_record_dict)

    # Format output request data
    request_data = {
        "prediction_id": prediction_id,
        "data": formated_list,
    }

    logger.info(f'request_data: {repr(request_data)}')

    request = requests.post(settings.HYDROML_URL, json=request_data)

    if request.status_code != 200:
        logger.error(f'Error on predict data via HydroML, {request.status_code}')
        return

    formated_response = []
    response = json.loads(request.json())

    # Format predicted values
    for record in response:
        try:
            result = result_mapping[str(record['prediction'])]

            formated_response.append({
                "datetime": dateutil.parser.isoparse(record['datetime']),
                "target_station_id": target_station_id,
                "variable_id": variable_id,
                "result": result,
            })
        except KeyError as e:
            logger.error(
                f'Error on predict_data for prediction "{prediction_id}": Invalid mapping for result "{record["prediction"]}".')
            raise Exception(e)

    # Update records' labels
    try:
        with conn.cursor() as cursor:
            cursor.executemany(f"""
                UPDATE raw_data 
                SET ml_flag = %(result)s
                WHERE station_id = %(target_station_id)s
                  AND variable_id = %(variable_id)s 
                  AND datetime = %(datetime)s;
            """, formated_response)
        conn.commit()
    except Exception as e:
        logger.error(f'Error on update raw_data: {repr(e)}')

@shared_task
def predict_preciptation_data():
    hydroml_params = HydroMLPredictionStation.objects.all()

    end_datetime = datetime.utcnow()
    start_datetime = end_datetime - timedelta(hours=2, minutes=30)

    for hydroml_param in hydroml_params:
        current_prediction = hydroml_param.prediction
        logger.info(f"Processing Prediction: {current_prediction.name}")
        station_ids = tuple(hydroml_param.neighborhood.neighborhood_stations.all().values_list('station_id', flat=True))

        result_mapping = {}
        mappings = HydroMLPredictionMapping.objects.filter(hydroml_prediction_id=hydroml_param.id)
        for mapping in mappings:
            result_mapping[mapping.prediction_result] = mapping.quality_flag.id

        try:
            predict_data(start_datetime=start_datetime,
                         end_datetime=end_datetime,
                         prediction_id=current_prediction.hydroml_prediction_id,
                         station_ids=station_ids,
                         target_station_id=hydroml_param.target_station.id,
                         variable_id=current_prediction.variable_id,
                         data_period_in_minutes=hydroml_param.data_period_in_minutes,
                         interval_in_minutes=hydroml_param.interval_in_minutes,
                         result_mapping=result_mapping)
        except Exception as e:
            logger.error(f'Error on predict_preciptation_data for "{current_prediction.name}": {repr(e)}')

################################################################

@shared_task
def process_hfdata_summary_tasks():
    # process only 500 HF summaries per execution
    unprocessed_hf_summaries = HFSummaryTask.objects.filter(started_at=None).values_list('station', 'variable', 'start_datetime', 'end_datetime').distinct()[:501]
    for  station_id, variable_id, s_datetime, e_datetime in unprocessed_hf_summaries:

        hf_summary_task = HFSummaryTask.objects.filter(started_at=None, station_id=station_id, variable_id=variable_id, start_datetime=s_datetime, end_datetime=e_datetime)
        hf_summary_task_ids = list(hf_summary_task.values_list('id', flat=True))

        # Updating Hourly summary task
        hfdata_summary(hf_summary_task_ids, station_id, variable_id, s_datetime, e_datetime)

def hfdata_summary(hf_summary_task_ids, station_id, variable_id, s_datetime, e_datetime):
    try:
        HFSummaryTask.objects.filter(id__in=hf_summary_task_ids).update(started_at=datetime.now(tz=pytz.UTC))
        calculate_hfdata_summary(station_id, variable_id, s_datetime, e_datetime)
    except Exception as err:
        logger.error('Error calculation hfdata summary for variable "{0}" and range ("{1}","{2}"). '.format(variable_id, s_datetime, s_datetime) + repr(err))
        db_logger.error('Error calculation hfdata summary for variable "{0}" and range ("{1}","{2}"). '.format(variable_id, s_datetime, s_datetime) + repr(err))
    else:
        HFSummaryTask.objects.filter(id__in=hf_summary_task_ids).update(finished_at=datetime.now(tz=pytz.UTC))

def calculate_hfdata_summary(station_id, variable_id, s_datetime, e_datetime):
    start_at = time()
    logger.info(f"HighFrequency summary started at {datetime.now(pytz.UTC)}")
    logger.info(f"HighFrequency summary parameters: {station_id} {variable_id} {s_datetime} {e_datetime}")

    _hfdata = HighFrequencyData.objects.filter(station_id=station_id, variable_id=variable_id, datetime__gte=s_datetime, datetime__lte=e_datetime)
    
    datetimes = list(_hfdata.values_list('datetime', flat=True))

    seconds = (max(datetimes)-min(datetimes)).total_seconds()+1
    interval = seconds/len(datetimes)

    data = list(_hfdata.values_list('measured', flat=True))

    _variable = Variable.objects.get(id=variable_id)
    if _variable.name == 'Sea Level':
        reads = process_wave_data(station_id, e_datetime, data, seconds)

    insert_rd(reads)

    logger.info(f'HighFrequency summary finished at {datetime.now(pytz.UTC)}. Took {time() - start_at} seconds.')

class wave(): # Wave object
    def __init__(self, frequency: float, height: float, phase_rad: float):
        self.height = height # Wave height in cm
        self.frequency = frequency # Frequency in Hz
        self.phase_rad = phase_rad # Phase is radians
        self.time = None
        self.wave = None

    def gen_sinewave(self, time):
        self.time = time
        self.wave = self.height*np.sin(self.frequency*2*np.pi*time+self.phase_rad)
        return self.wave

def fft_decompose(data, DEBUG = False):
    MEASUREMENT_PERIOD = 1 #1 Second

    MIN_AMPLITUDE = 0.01 #In m

    MIN_FREQUENCY = 0.0001 #In Hz
    MAX_FREQUENCY = 0.3 #In Hz

    # https://stackoverflow.com/questions/59725933/plot-fft-as-a-set-of-sine-waves-in-python
    t = np.arange(0, len(data)*MEASUREMENT_PERIOD, MEASUREMENT_PERIOD)

    fft = np.fft.fft(data)
    fftfreq = np.fft.fftfreq(len(t), MEASUREMENT_PERIOD)

    wave_list = []
    sinewave_list = []
    
    mid = len(t)//2 + 1    
    for i in range(mid):
        phase_rad = cmath.phase(fft[i])+np.pi/2
        amplitude = 2*abs(fft[i])/len(t)
        frequency = fftfreq[i]

        if MIN_FREQUENCY <= frequency <= MAX_FREQUENCY and MIN_AMPLITUDE <= amplitude:
            W = wave(frequency=frequency, height=amplitude, phase_rad=phase_rad)

            wave_list.append(W)
            sinewave_list.append(W.gen_sinewave(t))

            if DEBUG:
                logger.info(f"WAVE  {len(sinewave_list)}:")
                logger.info(f"Height: {amplitude}:")
                logger.info(f"Frequency: {frequency}:")
                logger.info(f"Phase: {math.degrees(phase_rad)}:")
    return wave_list

def get_top_wave_components(wave_list):
    # Top 5 amp
    wave_list.sort(key=lambda W: abs(W.height), reverse=True)
    top_wave_components = wave_list[:5]
    return top_wave_components

def process_wave_data(station_id, e_datetime, data, seconds):
    avg = np.mean(data)
    std = np.std(data)

    reads = [['Sea Level [MIN]', np.min(data)],
             ['Sea Level [MAX]', np.max(data)],
             ['Sea Level [AVG]', avg],
             ['Sea Level [STDV]', std],
             ['Significant Wave Height', 4*std]]

    # Normalizing Sea Level by average
    norm_data = data-avg

    wave_list = fft_decompose(norm_data, DEBUG = True)
    wave_list = get_top_wave_components(wave_list)

    for i, W in enumerate(wave_list):
        name = 'Wave Component '+str(i+1)
        reads += [[name+' Frequency', W.frequency],
                  [name+' Amplitude', W.height],
                  [name+' Phase', math.degrees(W.phase_rad) % 360]]

    for read in reads:
        read[0] = Variable.objects.get(name=read[0]).id

    now = datetime.now()

    df = pd.DataFrame(reads, columns=['variable_id', 'measured'])
    df['station_id'] = station_id
    df['datetime'] = e_datetime
    df['updated_at'] = now
    df['created_at'] = now
    df['seconds'] = seconds
    df['is_daily'] = False
    df['quality_flag'] = None
    df['qc_range_quality_flag'] = None
    df['qc_range_description'] = None
    df['qc_step_quality_flag'] = None
    df['qc_step_description'] = None
    df['qc_persist_quality_flag'] = None
    df['qc_persist_description'] = None
    df['manual_flag'] = None
    df['consisted'] = None

    cols = ["station_id", "variable_id", "seconds", "datetime", "measured", "quality_flag", "qc_range_quality_flag",
           "qc_range_description", "qc_step_quality_flag", "qc_step_description", "qc_persist_quality_flag",
           "qc_persist_description", "manual_flag", "consisted", "is_daily"]

    reads = df[cols].values.tolist()

    return reads




######################################################################
# logic to submit station information to OSCAR given a request
@shared_task
def export_station_to_oscar(request):
    # get wigos ID's
    selected_ids = request.POST.getlist('selected_ids[]')

    api_token = request.POST.get('api_token')

    # Log the wigos id
    print(f"Selected IDs: {selected_ids}")
    # print(f"Token entered: {api_token}")

    if selected_ids:

        stations = Station.objects.filter(wigos__in=selected_ids)

        oscar_status_msg_list = []

        try:
            for obj in stations:
                station_info = []
                
                station_info.append(str(obj.wigos))
                station_info.append(str(obj.name))
                station_info.append(str(datetime.strptime(f'{obj.begin_date}', '%Y-%m-%d %H:%M:%S%z').strftime('%Y-%m-%dT%H:%M:%SZ')))
                station_info.append(str(obj.latitude))
                station_info.append(str(obj.longitude))
                station_info.append(str(obj.elevation))
                station_info.append(str(obj.wmo_region.notation))
                station_info.append(str(obj.wmo_station_type.notation))
                station_info.append(str(obj.reporting_status.notation))
                station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')))
                station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dZ')))
                station_info.append(str(obj.country.notation))

                # export station information to OSCAR
                oscar_status_msg_list.append(exso.surface_to_oscar(station_info, api_token=api_token))

        except Exception as e:
            oscar_status_msg_list.append({'code': 406, 'description': f'{e}'})
        
    else:
        oscar_status_msg_list.append({'code': 412, 'description': 'No WIGOS ID was provided'})

    return  oscar_status_msg_list


# logic to submit station information to OSCAR given the wigos id and the api token
@shared_task
def export_station_to_oscar_wigos(selected_ids, api_token, cleaned_data):

    # Log the wigos id
    print(f"Selected IDs: {selected_ids[0]}")
    # print(f"Token entered: {api_token}")

    if selected_ids[0]:

        try:
            station_info = []
            
            station_info.append(str(selected_ids[0]))
            station_info.append(str(cleaned_data['name']))
            station_info.append(str(datetime.strptime(f"{cleaned_data['begin_date']}", '%Y-%m-%d %H:%M:%S%z').strftime('%Y-%m-%dT%H:%M:%SZ')))
            station_info.append(str(cleaned_data['latitude']))
            station_info.append(str(cleaned_data['longitude']))
            station_info.append(str(cleaned_data['elevation']))
            station_info.append(str(WMORegion.objects.filter(name=cleaned_data['wmo_region']).values_list('notation', flat=True).first()))
            station_info.append(str(WMOStationType.objects.filter(name=cleaned_data['wmo_station_type']).values_list('notation', flat=True).first()))
            station_info.append(str(WMOReportingStatus.objects.filter(name=cleaned_data['reporting_status']).values_list('notation', flat=True).first()))
            station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')))
            station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dZ')))
            station_info.append(str(Country.objects.filter(name=cleaned_data['country']).values_list('notation', flat=True).first()))

            # export station information to OSCAR
            return exso.surface_to_oscar(station_info, api_token=api_token)

        except Exception as e:
            return {'code': 406, 'description': f'{e}'}
        
    else:
        return {'code': 412, 'description': 'No WIGOS ID was provided'}


# MANUAL STATIONS ONLY** Tasks the looks for stations that are set for international exchange (see STATION model) and transmits them to wis2box
def manual_transmit_wis2box(id, station_name, regional_transmit, local_transmit):
    wis2_logs = []  # Stores log messages for debugging and reporting
    wis2_message = []
    regional_success = False
    local_success = False

    # Get UTC time and truncate to the hour
    unformated_now = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # Format it as 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    formated_now = unformated_now.strftime('%Y-%m-%d %H:%M:%S+00')

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        wis2_logs.append(formatted_message)  # Store it in wis2_logs
        

    logging.info(f'MWS Transmission: Processing data for Station: {station_name}')
    log_message(f'MWS Transmission: Processing data for Station: {station_name}')

    # Initialize data_row dictionary with default None values for each weather parameter
    data_row = {key: None for key in [
        'wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 'wmo_block_number',
        'wmo_station_number', 'station_type', 'year', 'month', 'day', 'hour', 'minute',
        'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl',
        'thermometer_height', 'anemometer_height', 'time_period_of_wind', 'rain_sensor_height', 
        'wind_indicator_manual', 'precipitation_indicator_manual', 'station_operation_indicator_manual',
        'lowest_cloud_height_manual', 'visibility_KM_manual', 'cloud_cover_tot_manual', 'wind_direction_manual',
        'wind_speed_manual', 'air_temperature_dry_bulb_manual', 'dew_point_temperature_manual',
        'air_temperature_wet_bulb_manual', 'relative_humidity_manual', 'pressure_at_station_level_manual',
        'pressure_at_sea_level_hpa_manual', 'precipitation_since_last_report_manual', 'precipitation_period_duration_manual',
        '24_hour_barometric_change_manual', '24_hour_rainfall_manual', 'present_weather_manual',
        'past_weather_type_1_manual', 'past_weather_type_2_manual', 'observed_cloud_coverage_manual',
        'low_cloud_type_manual', 'middle_cloud_type_manual', 'high_cloud_type_manual', 'state_of_sky_manual',
        'direction_of_cl_clouds_manual', 'direction_of_cm_clouds_manual', 'direction_of_ch_clouds_manual',
        'air_temperature_dry_bulb_max_manual', 'cloud_coverage_level_1_manual', 'genus_of_cloud_level_1_manual',
        'height_of_cloud_base_level_1_manual', 'cloud_coverage_level_2_manual', 'genus_of_cloud_level_2_manual',
        'height_of_cloud_base_level_2_manual', 'cloud_coverage_level_3_manual', 'genus_of_cloud_level_3_manual',
        'height_of_cloud_base_level_3_manual', 'cloud_coverage_level_4_manual', 'genus_of_cloud_level_4_manual',
        'height_of_cloud_base_level_4_manual', 'special_phenomenon_manual']}

    # Helper function to execute database queries safely
    def execute_query(query, params, error_message):
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                return [list(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f'{error_message}: {e}')
            log_message(f'{error_message}: {e}')
            return []


    ##########################################################################################
    # Fetch station information from the database
    station_query = """
        SELECT wigos, latitude, longitude, elevation, is_automatic
        FROM wx_station WHERE id = %s
    """
    station_data = execute_query(station_query, [id], "Error fetching station info")
    if station_data:
        row = station_data[0]
        wsi_parts = row[0].split("-") if row[0] else [None] * 4
        data_row.update({
            'wsi_series': wsi_parts[0], 
            'wsi_issuer': wsi_parts[1],
            'wsi_issue_number': wsi_parts[2], 
            'wsi_local': wsi_parts[3],
            'latitude': round(row[1], 5), 
            'longitude': round(row[2], 5),
            'station_height_above_msl': row[3], 
            'barometer_height_above_msl': round(row[3] + 1.7, 2), # 1.7 metre above station_height_above_msl
            # the settings below will be hard coded until further notice
            'thermometer_height': 1.7, 
            'anemometer_height': 10, 
            'time_period_of_wind': -5,
            'rain_sensor_height': 1.5, 
            'station_type': 0
        })

    # Populate year, month, day, hour, minute in data_row
    if unformated_now:
        # Add the extracted date components to data_row
        data_row['year'] = unformated_now.year
        data_row['month'] = unformated_now.month
        data_row['day'] = unformated_now.day
        data_row['hour'] = unformated_now.hour
        data_row['minute'] = unformated_now.minute

    ##########################################################################################

    # ------------- GETTING MANUAL STATIONS DATA START ------------- #
    # query to get the measure values
    manual_data_query = """
        SELECT variable_id, measured
        FROM raw_data
        WHERE station_id = %s 
        AND datetime = %s
        AND variable_id IN (10, 16, 18, 19, 30, 50, 55, 60, 61,
                            4048, 4050, 4055, 4056, 4057);
    """

    # query to get the code values
    manual_data_query_code = """
        SELECT variable_id, code
        FROM raw_data
        WHERE station_id = %s 
        AND datetime = %s
        AND variable_id IN (1001, 1002, 1003, 1004, 1006, 1007, 1008, 1009, 1010, 1011,
                            1012, 1013, 1014, 1015, 1016, 1017, 1018, 1019, 1020, 4005,
                            4006, 4040, 4041, 4042, 4044, 4045, 4046, 4047, 4043);
    """

    # Execute the query
    secondary_station_raw_data = execute_query(manual_data_query, [id, formated_now], "Error fetching raw data")
    secondary_station_raw_data_code = execute_query(manual_data_query_code, [id, formated_now], "Error fetching raw data")

    # Mapping of variable_id to dictionary keys
    variable_mapping = {
        4040: 'wind_indicator_manual', 4041: 'precipitation_indicator_manual', 4042: 'station_operation_indicator_manual', 
        4048: 'lowest_cloud_height_manual', 4056: 'visibility_KM_manual', 1001: 'cloud_cover_tot_manual', 
        55: 'wind_direction_manual', 50: 'wind_speed_manual', 10: 'air_temperature_dry_bulb_manual', 
        19: 'dew_point_temperature_manual', 18: 'air_temperature_wet_bulb_manual', 30: 'relative_humidity_manual', 
        60: 'pressure_at_station_level_manual', 61: 'pressure_at_sea_level_hpa_manual', 4050: 'precipitation_since_last_report_manual', 
        4043: 'precipitation_period_duration_manual', 4057: '24_hour_barometric_change_manual', 4055: '24_hour_rainfall_manual', 
        1002: 'present_weather_manual', 1003: 'past_weather_type_1_manual', 1004: 'past_weather_type_2_manual', 
        4005: 'observed_cloud_coverage_manual', 1006: 'low_cloud_type_manual', 1007: 'middle_cloud_type_manual', 
        1008: 'high_cloud_type_manual', 4044: 'state_of_sky_manual', 4045: 'direction_of_cl_clouds_manual', 
        4046: 'direction_of_cm_clouds_manual', 4047: 'direction_of_ch_clouds_manual', 16: 'air_temperature_dry_bulb_max_manual', 
        1009: 'cloud_coverage_level_1_manual', 1010: 'genus_of_cloud_level_1_manual', 1011: 'height_of_cloud_base_level_1_manual', 
        1012: 'cloud_coverage_level_2_manual', 1013: 'genus_of_cloud_level_2_manual', 1014: 'height_of_cloud_base_level_2_manual', 
        1015: 'cloud_coverage_level_3_manual', 1016: 'genus_of_cloud_level_3_manual', 1017: 'height_of_cloud_base_level_3_manual', 
        1018: 'cloud_coverage_level_4_manual', 1019: 'genus_of_cloud_level_4_manual', 1020: 'height_of_cloud_base_level_4_manual', 
        4006: 'special_phenomenon_manual'}


    # Populate the dictionary with retrieved data
    for row in secondary_station_raw_data:
        variable_id, measured_value = row
        if variable_id in variable_mapping:
            data_row[variable_mapping[variable_id]] = measured_value   

    # Populate the dictionary with retrieved data (only data entries which store codes)
    for row in secondary_station_raw_data_code:
        variable_id, code_value = row
        if variable_id in variable_mapping:
            data_row[variable_mapping[variable_id]] = code_value   
    # ------------- GETTING MANUAL STATIONS DATA END ------------- #


    # Helper function to upload CSV to MinIO storage
    def upload_to_minio(credentials, label, path):
        try:
            if label == "regional":
                client = Minio(
                    endpoint=f"{credentials.regional_wis2_ip_address}:{credentials.regional_wis2_port}",
                    access_key=credentials.regional_wis2_username,
                    secret_key=str(credentials.get_regional_password()),
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
                
            else:
                client = Minio(
                    endpoint=f"{credentials.local_wis2_ip_address}:{credentials.local_wis2_port}",
                    access_key=credentials.local_wis2_username,
                    secret_key=str(credentials.get_local_password()), 
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
            
            # if add_gts:
            #     # Function to set tag for file based on hour
            #     def set_tag(hour, minute):
            #         if hour % 6 == 0 and minute == 0:
            #             return 'main'
            #         elif hour % 3 == 0 and minute == 0:
            #             return 'intermediate'
            #         elif minute == 0:
            #             return 'non-standard'
            #         else:
            #             return 'other'
                    
            #     # Now get the tag
            #     tag = set_tag(data_row['hour'], data_row['minute'])

            #     # setting the base stations id as the id used in the file name
            #     if base_aws:
            #         filename = f"wmo_data_{id}_{tag}.csv"
            #     else:
            #         filename = f"wmo_data_{secondary_station}_{tag}.csv"
            # else:
            #     filename = f'wmo_data_{id}.csv'
            filename = f'wmo_data_{id}.csv'

            # Note that 'settings.WIS2BOX_TOPIC_HIERARCHY + filename' simply just tacks on the filename no '/' separation
            # see 'https://docs.wis2box.wis.wmo.int/en/latest/user/data-ingest.html' for and example
            # client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + filename, temp_csv_path)
            # Note: The filename and the path are not the same, but i don't believe this affects wis2 reception if an error occurs consider 
            # having the names match
            client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + '/' + filename, path)

            # Add the file name
            wis2_message.append(filename)

            logging.info(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            log_message(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            return True
        except Exception as e:
            logging.error(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            log_message(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            return False

    try:
        # Save the data into a temporary CSV file
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w+', delete=False) as temp_csv:
            csv_writer = csv.DictWriter(temp_csv, fieldnames=data_row.keys())
            csv_writer.writeheader()
            csv_writer.writerow(data_row)
            temp_csv_path = temp_csv.name

            # Capture the CSV content into a string
            temp_csv.seek(0)  # Move to the beginning of the file
            csv_content = temp_csv.read()

        # Transmit data to regional and local WIS2Box if required
        if regional_transmit:
            regional_success = upload_to_minio(RegionalWisCredentials.objects.first(), "regional", temp_csv_path)
            if regional_success:
                log_message('Push to the Regional WIS2BOX was successfull')
            else:
                log_message('Push to the Regional WIS2BOX failed')
        if local_transmit:
            local_success = upload_to_minio(LocalWisCredentials.objects.first(), "local", temp_csv_path)
            if local_success:
                log_message('Push to the Local WIS2BOX was successfull')
            else:
                log_message('Push to the Local WIS2BOX failed')

        # Add the CSV data to the list
        wis2_message.append(csv_content)
    finally:
        # Clean up the temporary CSV file
        os.remove(temp_csv_path)
        logging.info(f'AWS Transmission for Station: {station_name} complete')
        log_message(f'AWS Transmission for Station: {station_name} complete')

        # if both regional and local transmission are selected both pushes must be success to get a 'success status'
        if regional_transmit and local_transmit:
            success_push = regional_success and local_success
        # else if only regional transmission is set the 'success status' is based on only regional transmission
        elif regional_transmit:
            success_push = regional_success
        # else if only local transmission is set the 'success status' is based on only local transmission
        elif local_transmit:
            success_push = local_success
        else:
            success_push = False

    return wis2_logs, wis2_message, success_push


# AWS STATIONS ONLY** Tasks the looks for stations that are set for international exchange (see STATION model) and transmits them to wis2box
def aws_transmit_wis2box(id, station_name, regional_transmit, local_transmit):
    wis2_logs = []  # Stores log messages for debugging and reporting
    wis2_message = []
    regional_success = False
    local_success = False

    # Get UTC time and truncate to the hour
    unformated_now = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # Format it as 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    formated_now = unformated_now.strftime('%Y-%m-%d %H:%M:%S+00')

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        wis2_logs.append(formatted_message)  # Store it in wis2_logs

    logging.info(f'AWS Transmission: Processing data for Station: {station_name}')
    log_message(f'AWS Transmission: Processing data for Station: {station_name}')

    # Initialize data_row dictionary with default None values for each weather parameter
    data_row = {key: None for key in [
        'wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 'wmo_block_number',
        'wmo_station_number', 'station_type', 'year', 'month', 'day', 'hour', 'minute',
        'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl',
        'station_pressure', 'msl_pressure', 'geopotential_height', 'thermometer_height',
        'air_temperature', 'dewpoint_temperature', 'relative_humidity',
        'method_of_ground_state_measurement', 'ground_state', 'method_of_snow_depth_measurement',
        'snow_depth', 'precipitation_intensity', 'anemometer_height', 'time_period_of_wind',
        'wind_direction', 'wind_speed', 'maximum_wind_gust_direction_10_minutes',
        'maximum_wind_gust_speed_10_minutes', 'maximum_wind_gust_direction_1_hour',
        'maximum_wind_gust_speed_1_hour', 'maximum_wind_gust_direction_3_hours',
        'maximum_wind_gust_speed_3_hours', 'rain_sensor_height', 'total_precipitation_1_hour',
        'total_precipitation_3_hours', 'total_precipitation_6_hours',
        'total_precipitation_12_hours', 'total_precipitation_24_hours']}

    # Helper function to execute database queries safely
    def execute_query(query, params, error_message):
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                return [list(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f'{error_message}: {e}')
            log_message(f'{error_message}: {e}')
            return []


    ##########################################################################################
    # Fetch station information from the database
    station_query = """
        SELECT wigos, latitude, longitude, elevation, is_automatic
        FROM wx_station WHERE id = %s
    """
    station_data = execute_query(station_query, [id], "Error fetching station info")
    if station_data:
        row = station_data[0]
        wsi_parts = row[0].split("-") if row[0] else [None] * 4
        data_row.update({
            'wsi_series': wsi_parts[0], 
            'wsi_issuer': wsi_parts[1],
            'wsi_issue_number': wsi_parts[2], 
            'wsi_local': wsi_parts[3],
            'latitude': round(row[1], 5), 
            'longitude': round(row[2], 5),
            'station_height_above_msl': row[3], 
            'barometer_height_above_msl': round(row[3] + 1.7, 2), # 1.7 metre above station_height_above_msl
            # the settings below will be hard coded until further notice
            'thermometer_height': 1.7, 
            'anemometer_height': 10, 
            'time_period_of_wind': -5,
            'rain_sensor_height': 1.5, 
            'station_type': 0
        })

    ##########################################################################################
    # Fetch the latest raw data measurements for specific variables
    raw_data_query = """
        SELECT variable_id, measured, datetime FROM raw_data
        WHERE station_id = %s AND datetime = %s
        AND variable_id IN (10, 19, 30, 51, 56, 60, 61);
    """
    raw_data = execute_query(raw_data_query, [id, formated_now], "Error fetching raw data")

    for row in raw_data:
        key, value = {
            10: ('air_temperature', round(row[1] + 273.15, 2)),
            19: ('dewpoint_temperature', round(row[1] + 273.15, 2)),
            30: ('relative_humidity', round(row[1])),
            51: ('wind_speed', round(row[1], 1)),
            56: ('wind_direction', int(row[1])),
            60: ('station_pressure', row[1] * 100),
            61: ('msl_pressure', row[1] * 100)
        }.get(row[0], (None, None))

        if key:  # Only update if a valid key was found
            data_row[key] = value
    
    # Populate year, month, day, hour, minute in data_row
    if unformated_now:
        # Add the extracted date components to data_row
        data_row['year'] = unformated_now.year
        data_row['month'] = unformated_now.month
        data_row['day'] = unformated_now.day
        data_row['hour'] = unformated_now.hour
        data_row['minute'] = unformated_now.minute


    ##########################################################################################
    # Fetch precipitation data for the past 24 hours
    precipitation_query = """
        SELECT sum_value 
        FROM hourly_summary 
        WHERE station_id = %s 
        AND variable_id = 0 
        AND datetime BETWEEN %s - interval '24 hours' AND %s
        ORDER BY datetime DESC;
    """
    precipitation_data = execute_query(precipitation_query, [id, formated_now, formated_now], "Error fetching precipitation data")
    precip_values = [row[0] for row in precipitation_data if row]
    if precip_values:

        if precip_values[3]:
            data_row.update({
                'total_precipitation_3_hours': round(sum(precip_values[:3]), 1)
            })
        elif precip_values[6]:
            data_row.update({
                'total_precipitation_6_hours': round(sum(precip_values[:6]), 1)
            })
        elif precip_values[12]:
            data_row.update({
                'total_precipitation_12_hours': round(sum(precip_values[:12]), 1)
            })

        data_row.update({
            'total_precipitation_1_hour': round(precip_values[0], 1),
            'total_precipitation_24_hours': round(sum(precip_values), 1)
        })


    ##########################################################################################
    # fetch wind SPEED 1 hour and 3 hour data
    wind_query = """
        SELECT max_value 
        FROM hourly_summary hs 
        WHERE station_id = %s 
        AND variable_id = 51 
        AND datetime BETWEEN %s - interval '3 hours' AND %s
        ORDER BY datetime DESC;
    """
    wind_data = execute_query(wind_query, [id, formated_now, formated_now], "Error fetching wind data")
    wind_hourly = [rows[0] for rows in wind_data if row]
    if wind_hourly:
        data_row['maximum_wind_gust_speed_1_hour'] = round(wind_hourly[0], 1)
        # Find the maximum wind gust speed in the 3-hour window and then round maximum value
        data_row['maximum_wind_gust_speed_3_hours'] = round(max(wind_hourly), 1)

    # fetch wind DIRECTION 1 hour and 3 hour data
    wind_direction_query = """
        SELECT max_value 
        FROM hourly_summary hs 
        WHERE station_id = %s 
        AND variable_id = 56 
        AND datetime BETWEEN %s - interval '3 hours' AND %s
        ORDER BY datetime DESC;
    """
    wind_direction_data = execute_query(wind_direction_query, [id, formated_now, formated_now], "Error fetching wind data")
    wind_direction_hourly = [rows[0] for rows in wind_direction_data if row]
    if wind_direction_hourly:
        data_row['maximum_wind_gust_direction_1_hour'] = int(wind_direction_hourly[0])
        # Find the maximum wind gust direction in the 3-hour window and then round maximum value
        data_row['maximum_wind_gust_direction_3_hours'] = int(max(wind_direction_hourly))
    ##########################################################################################


    # Helper function to upload CSV to MinIO storage
    def upload_to_minio(credentials, label, path):
        try:
            if label == "regional":
                client = Minio(
                    endpoint=f"{credentials.regional_wis2_ip_address}:{credentials.regional_wis2_port}",
                    access_key=credentials.regional_wis2_username,
                    secret_key=str(credentials.get_regional_password()),
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
                
            else:
                client = Minio(
                    endpoint=f"{credentials.local_wis2_ip_address}:{credentials.local_wis2_port}",
                    access_key=credentials.local_wis2_username,
                    secret_key=str(credentials.get_local_password()), 
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
                
            filename = f'wmo_data_{id}.csv'

            # Note that 'settings.WIS2BOX_TOPIC_HIERARCHY + filename' simply just tacks on the filename no '/' separation
            # see 'https://docs.wis2box.wis.wmo.int/en/latest/user/data-ingest.html' for and example
            # client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + filename, temp_csv_path)
            client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + '/' + filename, path)

            logging.info(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            log_message(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            return True
        except Exception as e:
            logging.error(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            log_message(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            return False

    try:
        # Save the data into a temporary CSV file
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w+', delete=False) as temp_csv:
            csv_writer = csv.DictWriter(temp_csv, fieldnames=data_row.keys())
            csv_writer.writeheader()
            csv_writer.writerow(data_row)
            temp_csv_path = temp_csv.name

            # Capture the CSV content into a string
            temp_csv.seek(0)  # Move to the beginning of the file
            csv_content = temp_csv.read()

        # Add the file path and CSV data to the list
        wis2_message.append(temp_csv_path)
        wis2_message.append(csv_content)

        # Transmit data to regional and local WIS2Box if required
        if regional_transmit:
            regional_success = upload_to_minio(RegionalWisCredentials.objects.first(), "regional", temp_csv_path)
            if regional_success:
                log_message('Push to the Regional WIS2BOX was successfull')
            else:
                log_message('Push to the Regional WIS2BOX failed')
        if local_transmit:
            local_success = upload_to_minio(LocalWisCredentials.objects.first(), "local", temp_csv_path)
            if local_success:
                log_message('Push to the Local WIS2BOX was successfull')
            else:
                log_message('Push to the Local WIS2BOX failed')
    finally:
        # Clean up the temporary CSV file
        os.remove(temp_csv_path)
        logging.info(f'AWS Transmission for Station: {station_name} complete')
        log_message(f'AWS Transmission for Station: {station_name} complete')
        
        # if both regional and local transmission are selected both pushes must be success to get a 'success status'
        if regional_transmit and local_transmit:
            success_push = regional_success and local_success
        # else if only regional transmission is set the 'success status' is based on only regional transmission
        elif regional_transmit:
            success_push = regional_success
        # else if only local transmission is set the 'success status' is based on only local transmission
        elif local_transmit:
            success_push = local_success
        else:
            success_push = False

    return wis2_logs, wis2_message, success_push


# transmit for transition stations
def transition_transmit_wis2box(id, station_name, regional_transmit, local_transmit, transit_station_id, base_aws, add_gts):
    wis2_logs = []  # Stores log messages for debugging and reporting
    wis2_message = []
    regional_success = False
    local_success = False

    # Get UTC time and truncate to the hour
    unformated_now = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # Format it as 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    formated_now = unformated_now.strftime('%Y-%m-%d %H:%M:%S+00')

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        wis2_logs.append(formatted_message)  # Store it in wis2_logs

    # setting which station to get AWS data from depending on the base_aws data
    # see the Wis2BoxPublish model for more information
    if not base_aws:
        secondary_station = id
        id = transit_station_id
    else:
        secondary_station = transit_station_id
        

    logging.info(f'AWS Transmission: Processing data for Station: {station_name}')
    log_message(f'AWS Transmission: Processing data for Station: {station_name}')

    # Initialize data_row dictionary with default None values for each weather parameter
    data_row = {key: None for key in [
        'wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 'wmo_block_number',
        'wmo_station_number', 'station_type', 'year', 'month', 'day', 'hour', 'minute',
        'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl',
        'station_pressure', 'msl_pressure', 'geopotential_height', 'thermometer_height',
        'air_temperature', 'dewpoint_temperature', 'relative_humidity',
        'method_of_ground_state_measurement', 'ground_state', 'method_of_snow_depth_measurement',
        'snow_depth', 'precipitation_intensity', 'anemometer_height', 'time_period_of_wind',
        'wind_direction', 'wind_speed', 'maximum_wind_gust_direction_10_minutes',
        'maximum_wind_gust_speed_10_minutes', 'maximum_wind_gust_direction_1_hour',
        'maximum_wind_gust_speed_1_hour', 'maximum_wind_gust_direction_3_hours',
        'maximum_wind_gust_speed_3_hours', 'rain_sensor_height', 'total_precipitation_1_hour',
        'total_precipitation_3_hours', 'total_precipitation_6_hours',
        'total_precipitation_12_hours', 'total_precipitation_24_hours', 
        # manual station data
        'synop_key_entry_data', 'wind_indicator_manual', 'precipitation_indicator_manual', 'station_operation_indicator_manual',
        'lowest_cloud_height_manual', 'visibility_KM_manual', 'cloud_cover_tot_manual', 'wind_direction_manual',
        'wind_speed_manual', 'air_temperature_dry_bulb_manual', 'dew_point_temperature_manual',
        'air_temperature_wet_bulb_manual', 'relative_humidity_manual', 'pressure_at_station_level_manual',
        'pressure_at_sea_level_hpa_manual', 'precipitation_since_last_report_manual', 'precipitation_period_duration_manual',
        '24_hour_barometric_change_manual', '24_hour_rainfall_manual', 'present_weather_manual',
        'past_weather_type_1_manual', 'past_weather_type_2_manual', 'observed_cloud_coverage_manual',
        'low_cloud_type_manual', 'middle_cloud_type_manual', 'high_cloud_type_manual', 'state_of_sky_manual',
        'direction_of_cl_clouds_manual', 'direction_of_cm_clouds_manual', 'direction_of_ch_clouds_manual',
        'air_temperature_dry_bulb_max_manual', 'cloud_coverage_level_1_manual', 'genus_of_cloud_level_1_manual',
        'height_of_cloud_base_level_1_manual', 'cloud_coverage_level_2_manual', 'genus_of_cloud_level_2_manual',
        'height_of_cloud_base_level_2_manual', 'cloud_coverage_level_3_manual', 'genus_of_cloud_level_3_manual',
        'height_of_cloud_base_level_3_manual', 'cloud_coverage_level_4_manual', 'genus_of_cloud_level_4_manual',
        'height_of_cloud_base_level_4_manual', 'special_phenomenon_manual']}

    # Helper function to execute database queries safely
    def execute_query(query, params, error_message):
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                return [list(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f'{error_message}: {e}')
            log_message(f'{error_message}: {e}')
            return []

    # ------------- GETTING BASE STATIONS DATA START ------------- #
    # Fetch station information from the database
    station_query = """
        SELECT wigos, latitude, longitude, elevation, is_automatic
        FROM wx_station WHERE id = %s
    """
    station_data = execute_query(station_query, [id], "Error fetching station info")
    if station_data:
        row = station_data[0]
        wsi_parts = row[0].split("-") if row[0] else [None] * 4
        data_row.update({
            'wsi_series': wsi_parts[0], 
            'wsi_issuer': wsi_parts[1],
            'wsi_issue_number': wsi_parts[2], 
            'wsi_local': wsi_parts[3],
            'latitude': round(row[1], 5), 
            'longitude': round(row[2], 5),
            'station_height_above_msl': row[3], 
            'barometer_height_above_msl': round(row[3] + 1.7, 2), # 1.7 metre above station_height_above_msl
            # the settings below will be hard coded until further notice
            'thermometer_height': 1.7, 
            'anemometer_height': 10, 
            'time_period_of_wind': -5,
            'rain_sensor_height': 1.5, 
            'station_type': 0
        })

    ##########################################################################################
    # Fetch the latest raw data measurements for specific variables
    raw_data_query = """
        SELECT variable_id, measured, datetime FROM raw_data
        WHERE station_id = %s AND datetime = %s
        AND variable_id IN (10, 19, 30, 51, 56, 60, 61);
    """
    raw_data = execute_query(raw_data_query, [id, formated_now], "Error fetching raw data")

    for row in raw_data:
        key, value = {
            10: ('air_temperature', round(row[1] + 273.15, 2)),
            19: ('dewpoint_temperature', round(row[1] + 273.15, 2)),
            30: ('relative_humidity', round(row[1])),
            51: ('wind_speed', round(row[1], 1)),
            56: ('wind_direction', int(row[1])),
            60: ('station_pressure', row[1] * 100),
            61: ('msl_pressure', row[1] * 100)
        }.get(row[0], (None, None))

        if key:  # Only update if a valid key was found
            data_row[key] = value
    
    # Populate year, month, day, hour, minute in data_row
    if unformated_now:
        # Add the extracted date components to data_row
        data_row['year'] = unformated_now.year
        data_row['month'] = unformated_now.month
        data_row['day'] = unformated_now.day
        data_row['hour'] = unformated_now.hour
        data_row['minute'] = unformated_now.minute

    ##########################################################################################
    # Fetch precipitation data for the past 24 hours
    precipitation_query = """
        SELECT sum_value 
        FROM hourly_summary 
        WHERE station_id = %s 
        AND variable_id = 0 
        AND datetime BETWEEN %s - interval '24 hours' AND %s
        ORDER BY datetime DESC;
    """
    precipitation_data = execute_query(precipitation_query, [id, formated_now, formated_now], "Error fetching precipitation data")
    precip_values = [row[0] for row in precipitation_data if row]
    if precip_values:

        if precip_values[3]:
            data_row.update({
                'total_precipitation_3_hours': round(sum(precip_values[:3]), 1)
            })
        elif precip_values[6]:
            data_row.update({
                'total_precipitation_6_hours': round(sum(precip_values[:6]), 1)
            })
        elif precip_values[12]:
            data_row.update({
                'total_precipitation_12_hours': round(sum(precip_values[:12]), 1)
            })

        data_row.update({
            'total_precipitation_1_hour': round(precip_values[0], 1),
            'total_precipitation_24_hours': round(sum(precip_values), 1)
        })


    ##########################################################################################
    # fetch wind SPEED 1 hour and 3 hour data
    wind_query = """
        SELECT max_value 
        FROM hourly_summary hs 
        WHERE station_id = %s 
        AND variable_id = 51 
        AND datetime BETWEEN %s - interval '3 hours' AND %s
        ORDER BY datetime DESC;
    """

    wind_data = execute_query(wind_query, [id, formated_now, formated_now], "Error fetching wind data")
    wind_hourly = [rows[0] for rows in wind_data if row]
    if wind_hourly:
        data_row['maximum_wind_gust_speed_1_hour'] = round(wind_hourly[0], 1)
        # Find the maximum wind gust speed in the 3-hour window and then round maximum value
        data_row['maximum_wind_gust_speed_3_hours'] = round(max(wind_hourly), 1)

    # fetch wind DIRECTION 1 hour and 3 hour data
    wind_direction_query = """
        SELECT max_value 
        FROM hourly_summary hs 
        WHERE station_id = %s 
        AND variable_id = 56 
        AND datetime BETWEEN %s - interval '3 hours' AND %s
        ORDER BY datetime DESC;
    """
    wind_direction_data = execute_query(wind_direction_query, [id, formated_now, formated_now], "Error fetching wind data")
    wind_direction_hourly = [rows[0] for rows in wind_direction_data if row]
    if wind_direction_hourly:
        data_row['maximum_wind_gust_direction_1_hour'] = int(wind_direction_hourly[0])
        # Find the maximum wind gust direction in the 3-hour window and then round maximum value
        data_row['maximum_wind_gust_direction_3_hours'] = int(max(wind_direction_hourly))
    # ------------- GETTING BASE STATIONS DATA END ------------- #


    # ------------- GETTING SECONDARY STATIONS DATA START ------------- #
    # query to get the measure values
    manual_data_query = """
        SELECT variable_id, measured
        FROM raw_data
        WHERE station_id = %s 
        AND datetime = %s
        AND variable_id IN (10, 16, 18, 19, 30, 50, 55, 60, 61,
                            4048, 4050, 4055, 4056, 4057);
    """

    # query to get the code values
    manual_data_query_code = """
        SELECT variable_id, code
        FROM raw_data
        WHERE station_id = %s 
        AND datetime = %s
        AND variable_id IN (1001, 1002, 1003, 1004, 1006, 1007, 1008, 1009, 1010, 1011,
                            1012, 1013, 1014, 1015, 1016, 1017, 1018, 1019, 1020, 4005,
                            4006, 4040, 4041, 4042, 4044, 4045, 4046, 4047, 4043);
    """

    # Execute the query
    secondary_station_raw_data = execute_query(manual_data_query, [secondary_station, formated_now], "Error fetching raw data")
    secondary_station_raw_data_code = execute_query(manual_data_query_code, [secondary_station, formated_now], "Error fetching raw data")

    # Mapping of variable_id to dictionary keys
    variable_mapping = {
        4040: 'wind_indicator_manual', 4041: 'precipitation_indicator_manual', 4042: 'station_operation_indicator_manual', 
        4048: 'lowest_cloud_height_manual', 4056: 'visibility_KM_manual', 1001: 'cloud_cover_tot_manual', 
        55: 'wind_direction_manual', 50: 'wind_speed_manual', 10: 'air_temperature_dry_bulb_manual', 
        19: 'dew_point_temperature_manual', 18: 'air_temperature_wet_bulb_manual', 30: 'relative_humidity_manual', 
        60: 'pressure_at_station_level_manual', 61: 'pressure_at_sea_level_hpa_manual', 4050: 'precipitation_since_last_report_manual', 
        4043: 'precipitation_period_duration_manual', 4057: '24_hour_barometric_change_manual', 4055: '24_hour_rainfall_manual', 
        1002: 'present_weather_manual', 1003: 'past_weather_type_1_manual', 1004: 'past_weather_type_2_manual', 
        4005: 'observed_cloud_coverage_manual', 1006: 'low_cloud_type_manual', 1007: 'middle_cloud_type_manual', 
        1008: 'high_cloud_type_manual', 4044: 'state_of_sky_manual', 4045: 'direction_of_cl_clouds_manual', 
        4046: 'direction_of_cm_clouds_manual', 4047: 'direction_of_ch_clouds_manual', 16: 'air_temperature_dry_bulb_max_manual', 
        1009: 'cloud_coverage_level_1_manual', 1010: 'genus_of_cloud_level_1_manual', 1011: 'height_of_cloud_base_level_1_manual', 
        1012: 'cloud_coverage_level_2_manual', 1013: 'genus_of_cloud_level_2_manual', 1014: 'height_of_cloud_base_level_2_manual', 
        1015: 'cloud_coverage_level_3_manual', 1016: 'genus_of_cloud_level_3_manual', 1017: 'height_of_cloud_base_level_3_manual', 
        1018: 'cloud_coverage_level_4_manual', 1019: 'genus_of_cloud_level_4_manual', 1020: 'height_of_cloud_base_level_4_manual', 
        4006: 'special_phenomenon_manual'}


    # Populate the dictionary with retrieved data
    for row in secondary_station_raw_data:
        variable_id, measured_value = row
        if variable_id in variable_mapping:
            data_row[variable_mapping[variable_id]] = measured_value   

    # Populate the dictionary with retrieved data (only data entries which store codes)
    for row in secondary_station_raw_data_code:
        variable_id, code_value = row
        if variable_id in variable_mapping:
            data_row[variable_mapping[variable_id]] = code_value   
    # ------------- GETTING SECONDARY STATIONS DATA END ------------- #


    # Helper function to upload CSV to MinIO storage
    def upload_to_minio(credentials, label, path):
        try:
            if label == "regional":
                client = Minio(
                    endpoint=f"{credentials.regional_wis2_ip_address}:{credentials.regional_wis2_port}",
                    access_key=credentials.regional_wis2_username,
                    secret_key=str(credentials.get_regional_password()),
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
                
            else:
                client = Minio(
                    endpoint=f"{credentials.local_wis2_ip_address}:{credentials.local_wis2_port}",
                    access_key=credentials.local_wis2_username,
                    secret_key=str(credentials.get_local_password()), 
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
            
            if add_gts:
                # Function to set tag for file based on hour
                def set_tag(hour, minute):
                    if hour % 6 == 0 and minute == 0:
                        return 'main'
                    elif hour % 3 == 0 and minute == 0:
                        return 'intermediate'
                    elif minute == 0:
                        return 'non-standard'
                    else:
                        return 'other'
                    
                # Now get the tag
                tag = set_tag(data_row['hour'], data_row['minute'])

                # setting the base stations id as the id used in the file name
                if base_aws:
                    filename = f"wmo_data_{id}_{tag}.csv"
                else:
                    filename = f"wmo_data_{secondary_station}_{tag}.csv"
            else:
                filename = f'wmo_data_{id}.csv'

            # Note that 'settings.WIS2BOX_TOPIC_HIERARCHY + filename' simply just tacks on the filename no '/' separation
            # see 'https://docs.wis2box.wis.wmo.int/en/latest/user/data-ingest.html' for and example
            # client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + filename, temp_csv_path)
            # Note: The filename and the path are not the same, but i don't believe this affects wis2 reception if an error occurs consider 
            # having the names match
            client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + '/' + filename, path)

            # Add the file name
            wis2_message.append(filename)

            logging.info(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            log_message(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            return True
        except Exception as e:
            logging.error(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            log_message(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            return False

    try:
        # Save the data into a temporary CSV file
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w+', delete=False) as temp_csv:
            csv_writer = csv.DictWriter(temp_csv, fieldnames=data_row.keys())
            csv_writer.writeheader()
            csv_writer.writerow(data_row)
            temp_csv_path = temp_csv.name

            # Capture the CSV content into a string
            temp_csv.seek(0)  # Move to the beginning of the file
            csv_content = temp_csv.read()

        # Transmit data to regional and local WIS2Box if required
        if regional_transmit:
            regional_success = upload_to_minio(RegionalWisCredentials.objects.first(), "regional", temp_csv_path)
            if regional_success:
                log_message('Push to the Regional WIS2BOX was successfull')
            else:
                log_message('Push to the Regional WIS2BOX failed')
        if local_transmit:
            local_success = upload_to_minio(LocalWisCredentials.objects.first(), "local", temp_csv_path)
            if local_success:
                log_message('Push to the Local WIS2BOX was successfull')
            else:
                log_message('Push to the Local WIS2BOX failed')

        # Add the CSV data to the list
        wis2_message.append(csv_content)
    finally:
        # Clean up the temporary CSV file
        os.remove(temp_csv_path)
        logging.info(f'AWS Transmission for Station: {station_name} complete')
        log_message(f'AWS Transmission for Station: {station_name} complete')

        # if both regional and local transmission are selected both pushes must be success to get a 'success status'
        if regional_transmit and local_transmit:
            success_push = regional_success and local_success
        # else if only regional transmission is set the 'success status' is based on only regional transmission
        elif regional_transmit:
            success_push = regional_success
        # else if only local transmission is set the 'success status' is based on only local transmission
        elif local_transmit:
            success_push = local_success
        else:
            success_push = False

    return wis2_logs, wis2_message, success_push



@shared_task
def wis2publish_cleanup():
    # clean up logs older than 24 hours
    try:
        # Delete logs older than 24 hours
        time_threshold = now() - timedelta(hours=24)
        deleted_count, _ = Wis2BoxPublishLogs.objects.filter(created_at__lt=time_threshold).delete()

        logger.info(f"WIS2 Publish Cleanup tasks: Successfully deleted {deleted_count} logs older than 24 hours")
    except Exception as e:
        logger.error("WIS2 Publish Cleanup tasks: Failed to delete logs older than 24 hours")
        logger.exception(f"Error: {e}")

    # Update publish logs number and fail logs number
    try:
        # Annotate each Wis2BoxPublish record with the count of related success and fail logs
        wis2publish_updates = Wis2BoxPublish.objects.annotate(
            success_count=Count('wis2boxpublishlogs', filter=Q(wis2boxpublishlogs__success_log=True)),
            fail_count=Count('wis2boxpublishlogs', filter=Q(wis2boxpublishlogs__success_log=False))
        )

        update_count = 0  # Counter to track the number of successful updates

        # Iterate over each annotated Wis2BoxPublish record and update the success/fail counts
        for record in wis2publish_updates:
            Wis2BoxPublish.objects.filter(id=record.id).update(
                publish_success=record.success_count,  # Set the success count
                publish_fail=record.fail_count  # Set the fail count
            )
            update_count += 1  # Increment the count of updated records
        
        # Log the number of records successfully updated
        logger.info(f"WIS2 Publish Cleanup tasks: Successfully updated publish_fail and publish_success for {update_count} records in the wis2publish table")
    except Exception as e:
        # Log an error message if the update process fails
        logger.error("WIS2 Publish Cleanup tasks: Failed to update publish_fail and publish_success for some records")
        logger.exception(f"Error: {e}")  # Log the full exception traceback for debugging


@shared_task
def wis2publish_task():
    now = datetime.now(pytz.UTC)

    # retrieving wis2 tranmit station which are set to publishing
    _publishing_stations = Wis2BoxPublish.objects.filter(publishing=True)

    for _entry in _publishing_stations:
        # check if the cron schedule for the station matches the time now
        if croniter.match(_entry.publishing_offset.cron_schedule, now):
            # check if transition station or not
            if _entry.transition:
                wis2push_transition.delay(_entry.id, now)
            else:
                wis2push_regular.delay(_entry.id, now)


@shared_task
# for instant pushing
def wis2publish_task_now(station_id):
    try:
        now = datetime.now(pytz.UTC)

        # grabbing a single entry base on the id
        _entry = Wis2BoxPublish.objects.get(id=station_id)

        if _entry.transition:
            wis2push_transition(_entry.id, now)
        else:
            wis2push_regular(_entry.id, now)
    except Exception as e:
        raise e



# push to wis2 for non transition stations
@shared_task
def wis2push_regular(entry_id, now):
    # get the entry
    _entry = Wis2BoxPublish.objects.get(id=entry_id)

    # get station obj
    _station_obj = Station.objects.get(id=_entry.station.id)

    # differentiating between manual and automatic stations
    if _station_obj.is_automatic:
        # get the data from the transmission
        wis2_logs, wis2_message, success_push = aws_transmit_wis2box(_entry.station.id, _entry.station.name, _entry.regional_wis2, _entry.local_wis2)
    else:
        # get the data from the transmission
        wis2_logs, wis2_message, success_push = manual_transmit_wis2box(_entry.station.id, _entry.station.name, _entry.regional_wis2, _entry.local_wis2)

    message_exist = bool(wis2_message)

    # create the logs entry
    Wis2BoxPublishLogs.objects.create(
        created_at=now, 
        last_modified= datetime.now(tz=pytz.UTC),
        publish_station_id=_entry.id,
        success_log=success_push,
        log="\n\n".join(wis2_logs),
        wis2message_exist=message_exist,
        wis2message="\n".join(wis2_message)
    )

    # call clean up tasks
    wis2publish_cleanup() 


# push to wis2 for transition stations
@shared_task
def wis2push_transition(entry_id, now):
    # get the entry
    _entry = Wis2BoxPublish.objects.get(id=entry_id)

    # get data from the transmission
    wis2_logs, wis2_message, success_push = transition_transmit_wis2box(_entry.station.id, 
                                                                        _entry.station.name, 
                                                                        _entry.regional_wis2, 
                                                                        _entry.local_wis2, 
                                                                        _entry.transit_station.id,
                                                                        _entry.base_aws,
                                                                        _entry.add_gts)

    message_exist = bool(wis2_message)

    # create the logs entry
    Wis2BoxPublishLogs.objects.create(
        created_at=now, 
        last_modified= datetime.now(tz=pytz.UTC),
        publish_station_id=_entry.id,
        success_log=success_push,
        log="\n\n".join(wis2_logs),
        wis2message_exist=message_exist,
        wis2message="\n".join(wis2_message)
    )

    # call clean up tasks
    wis2publish_cleanup() 